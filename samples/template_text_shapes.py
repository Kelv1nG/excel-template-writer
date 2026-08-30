"""Generate maintained examples for editable template-authored text shapes."""

from __future__ import annotations

from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    BLUE,
    LIGHT_BLUE,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("template_text_shapes")
_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_SHEET = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_R = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"
_DRAWING_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
)
_DRAWING_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.drawing+xml"


def _qn(namespace: str, name: str) -> str:
    """Return one expanded ElementTree XML name.

    Args:
        namespace: XML namespace URI.
        name: Local element or attribute name.

    Returns:
        Expanded ``{namespace}name`` value.
    """

    return f"{{{namespace}}}{name}"


def _shape_anchor(
    *,
    row: int,
    column: int,
    identifier: int,
    text: str,
    preset: str,
    fill_color: str,
) -> ElementTree.Element:
    """Build one editable, styled, one-cell DrawingML text shape.

    Args:
        row: Zero-based anchor row.
        column: Zero-based anchor column.
        identifier: Non-visual object ID within the drawing part.
        text: Literal text stored in the shape.
        preset: DrawingML preset geometry name.
        fill_color: Six-digit RGB fill color.

    Returns:
        Complete one-cell anchor containing the authored shape.
    """

    anchor = ElementTree.Element(_qn(_XDR, "oneCellAnchor"))
    marker = ElementTree.SubElement(anchor, _qn(_XDR, "from"))
    ElementTree.SubElement(marker, _qn(_XDR, "col")).text = str(column)
    ElementTree.SubElement(marker, _qn(_XDR, "colOff")).text = "9525"
    ElementTree.SubElement(marker, _qn(_XDR, "row")).text = str(row)
    ElementTree.SubElement(marker, _qn(_XDR, "rowOff")).text = "19050"
    ElementTree.SubElement(
        anchor,
        _qn(_XDR, "ext"),
        {"cx": "2286000", "cy": "762000"},
    )
    shape = ElementTree.SubElement(anchor, _qn(_XDR, "sp"))
    non_visual = ElementTree.SubElement(shape, _qn(_XDR, "nvSpPr"))
    ElementTree.SubElement(
        non_visual,
        _qn(_XDR, "cNvPr"),
        {
            "id": str(identifier),
            "name": f"Sample text shape {identifier}",
            "descr": "Editable template-authored DrawingML text shape",
        },
    )
    ElementTree.SubElement(non_visual, _qn(_XDR, "cNvSpPr"), {"txBox": "1"})
    properties = ElementTree.SubElement(shape, _qn(_XDR, "spPr"))
    geometry = ElementTree.SubElement(
        properties,
        _qn(_A, "prstGeom"),
        {"prst": preset},
    )
    ElementTree.SubElement(geometry, _qn(_A, "avLst"))
    fill = ElementTree.SubElement(properties, _qn(_A, "solidFill"))
    ElementTree.SubElement(fill, _qn(_A, "srgbClr"), {"val": fill_color})
    line = ElementTree.SubElement(properties, _qn(_A, "ln"), {"w": "19050"})
    line_fill = ElementTree.SubElement(line, _qn(_A, "solidFill"))
    ElementTree.SubElement(line_fill, _qn(_A, "srgbClr"), {"val": "1F4E78"})

    text_body = ElementTree.SubElement(shape, _qn(_XDR, "txBody"))
    body = ElementTree.SubElement(
        text_body,
        _qn(_A, "bodyPr"),
        {
            "wrap": "square",
            "anchor": "ctr",
            "lIns": "91440",
            "rIns": "91440",
        },
    )
    ElementTree.SubElement(body, _qn(_A, "spAutoFit"))
    ElementTree.SubElement(text_body, _qn(_A, "lstStyle"))
    paragraph = ElementTree.SubElement(text_body, _qn(_A, "p"))
    run = ElementTree.SubElement(paragraph, _qn(_A, "r"))
    run_properties = ElementTree.SubElement(
        run,
        _qn(_A, "rPr"),
        {"lang": "en-US", "sz": "1200", "b": "1"},
    )
    run_fill = ElementTree.SubElement(run_properties, _qn(_A, "solidFill"))
    ElementTree.SubElement(run_fill, _qn(_A, "srgbClr"), {"val": "1F1F1F"})
    ElementTree.SubElement(run, _qn(_A, "t")).text = text
    ElementTree.SubElement(paragraph, _qn(_A, "endParaRPr"), {"lang": "en-US"})
    ElementTree.SubElement(
        anchor,
        _qn(_XDR, "clientData"),
        {"fLocksWithSheet": "1", "fPrintsWithSheet": "1"},
    )
    return anchor


def _inject_shape_drawings(
    path: Path,
    drawings: tuple[tuple[ElementTree.Element, ...], ...],
) -> Path:
    """Add relationship-free shape-only drawing parts to a saved workbook.

    Args:
        path: Saved XLSX package to augment in place through atomic replacement.
        drawings: One ordered shape-anchor collection per worksheet.

    Returns:
        Rewritten workbook path.
    """

    temporary_path = path.with_name(f"{path.stem}-drawingml.xlsx")
    with ZipFile(path, "r") as source, ZipFile(temporary_path, "w") as destination:
        source_names = frozenset(source.namelist())
        replacements: dict[str, bytes] = {}
        additions: dict[str, bytes] = {}
        content_types = ElementTree.fromstring(source.read("[Content_Types].xml"))

        for sheet_index, anchors in enumerate(drawings, start=1):
            drawing_part = f"xl/drawings/drawing{sheet_index}.xml"
            drawing_root = ElementTree.Element(_qn(_XDR, "wsDr"))
            drawing_root.extend(anchors)
            additions[drawing_part] = ElementTree.tostring(
                drawing_root,
                encoding="utf-8",
            )
            ElementTree.SubElement(
                content_types,
                _qn(_CONTENT_TYPES, "Override"),
                {
                    "PartName": f"/{drawing_part}",
                    "ContentType": _DRAWING_CONTENT_TYPE,
                },
            )

            relationships_part = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
            if relationships_part in source_names:
                relationships = ElementTree.fromstring(source.read(relationships_part))
            else:
                relationships = ElementTree.Element(_qn(_PACKAGE_R, "Relationships"))
            used_ids = {relationship.get("Id") for relationship in relationships}
            relationship_index = 1
            while f"rId{relationship_index}" in used_ids:
                relationship_index += 1
            relationship_id = f"rId{relationship_index}"
            ElementTree.SubElement(
                relationships,
                _qn(_PACKAGE_R, "Relationship"),
                {
                    "Id": relationship_id,
                    "Type": _DRAWING_RELATIONSHIP,
                    "Target": f"/{drawing_part}",
                },
            )
            relationship_xml = ElementTree.tostring(relationships, encoding="utf-8")
            if relationships_part in source_names:
                replacements[relationships_part] = relationship_xml
            else:
                additions[relationships_part] = relationship_xml

            worksheet_part = f"xl/worksheets/sheet{sheet_index}.xml"
            worksheet = ElementTree.fromstring(source.read(worksheet_part))
            ElementTree.SubElement(
                worksheet,
                _qn(_SHEET, "drawing"),
                {_qn(_R, "id"): relationship_id},
            )
            replacements[worksheet_part] = ElementTree.tostring(
                worksheet,
                encoding="utf-8",
            )

        replacements["[Content_Types].xml"] = ElementTree.tostring(
            content_types,
            encoding="utf-8",
        )
        for member in source.infolist():
            destination.writestr(
                member,
                replacements.get(member.filename, source.read(member.filename)),
            )
        for name, data in additions.items():
            destination.writestr(name, data)
    temporary_path.replace(path)
    return path


def _add_table(sheet: Worksheet, *, shift: str) -> None:
    """Add the table source that drives each shape movement example.

    Args:
        sheet: Worksheet receiving the source table.
        shift: Either the whole-row or isolated-cell shift policy.
    """

    sheet["A4"] = "Identifier"
    sheet["B4"] = "Plan"
    sheet["C4"] = "Actual"
    paint(sheet, "A4:C4", fill=BLUE, bold=True, font_color=WHITE, horizontal="center")
    shift_option = "" if shift == "rows" else ' shift="cells"'
    sheet["A5"] = f"{{% for item in items{shift_option} %}}{{{{ item.identifier }}}}"
    sheet["B5"] = "{{ item.plan }}"
    sheet["C5"] = "{{ item.actual }}{% endfor %}"
    paint(sheet, "A5:C5", fill=LIGHT_BLUE)
    sheet["B5"].number_format = "$#,##0"
    sheet["C5"].number_format = "$#,##0"


def _prepare_sheet(sheet: Worksheet, *, shift: str, description: str) -> None:
    """Prepare one visible static text-shape movement example.

    Args:
        sheet: Worksheet to initialize.
        shift: Table shift policy used on the sheet.
        description: Visible explanation of the expected movement.
    """

    prepare_sheet(
        sheet,
        "Template-authored editable text shapes",
        description,
        widths=(22, 15, 15, 3, 22, 14, 14, 14),
    )
    _add_table(sheet, shift=shift)


def _drawing_state(path: Path, part: str) -> list[tuple[int, int, str, str]]:
    """Read marker, preset geometry, and literal text from one drawing part.

    Args:
        path: XLSX package containing the drawing.
        part: Drawing XML filename below ``xl/drawings``.

    Returns:
        Zero-based row/column markers, geometry names, and literal shape text.
    """

    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(f"xl/drawings/{part}"))
    result: list[tuple[int, int, str, str]] = []
    for anchor in root:
        marker = anchor.find(_qn(_XDR, "from"))
        geometry = anchor.find(f".//{{{_A}}}prstGeom")
        assert marker is not None and geometry is not None
        row = marker.find(_qn(_XDR, "row"))
        column = marker.find(_qn(_XDR, "col"))
        assert row is not None and row.text is not None
        assert column is not None and column.text is not None
        text = "".join(element.text or "" for element in anchor.iter(_qn(_A, "t")))
        result.append((int(row.text), int(column.text), geometry.get("prst", ""), text))
    return result


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build whole-row and cell-lane templates with editable text shapes.

    Args:
        path: Destination for the generated template workbook.

    Returns:
        Verified template path containing two shape drawing parts.
    """

    workbook = Workbook()
    row_sheet = workbook.active
    row_sheet.title = "Row shift"
    _prepare_sheet(
        row_sheet,
        shift="rows",
        description="The E4 box stays fixed; the text callout at E16 moves to E27.",
    )
    lane_sheet = workbook.create_sheet("Cell lane")
    _prepare_sheet(
        lane_sheet,
        shift="cells",
        description="The A16 arrow moves to A27; the E16 box remains outside the A:C lane.",
    )
    saved = atomic_save(workbook, path)
    workbook.close()
    return _inject_shape_drawings(
        saved,
        (
            (
                _shape_anchor(
                    row=3,
                    column=4,
                    identifier=1,
                    text="Fixed decorative title box",
                    preset="roundRect",
                    fill_color="D9EAF7",
                ),
                _shape_anchor(
                    row=15,
                    column=4,
                    identifier=2,
                    text="Moves with the table; {{ shape.text }} remains literal",
                    preset="wedgeRectCallout",
                    fill_color="E2F0D9",
                ),
            ),
            (
                _shape_anchor(
                    row=15,
                    column=0,
                    identifier=1,
                    text="Inside A:C lane",
                    preset="rightArrow",
                    fill_color="E2F0D9",
                ),
                _shape_anchor(
                    row=15,
                    column=4,
                    identifier=2,
                    text="Outside lane",
                    preset="roundRect",
                    fill_color="D9EAF7",
                ),
            ),
        ),
    )


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render twelve rows and verify shape movement plus static rich text.

    Args:
        template_path: Authored template workbook to render.
        output_path: Separate destination for the rendered workbook.

    Returns:
        Verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {
            "items": [
                {
                    "identifier": f"Item {index:02d}",
                    "plan": 1_000 + index * 80,
                    "actual": 950 + index * 95,
                }
                for index in range(1, 13)
            ]
        },
    )
    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        for sheet_name in ("Row shift", "Cell lane"):
            sheet = workbook[sheet_name]
            assert [sheet[f"A{row}"].value for row in range(5, 17)] == [
                f"Item {index:02d}" for index in range(1, 13)
            ]
    finally:
        workbook.close()

    assert _drawing_state(template_path, "drawing1.xml") == [
        (3, 4, "roundRect", "Fixed decorative title box"),
        (
            15,
            4,
            "wedgeRectCallout",
            "Moves with the table; {{ shape.text }} remains literal",
        ),
    ]
    assert _drawing_state(output_path, "drawing1.xml") == [
        (3, 4, "roundRect", "Fixed decorative title box"),
        (
            26,
            4,
            "wedgeRectCallout",
            "Moves with the table; {{ shape.text }} remains literal",
        ),
    ]
    assert _drawing_state(template_path, "drawing2.xml") == [
        (15, 0, "rightArrow", "Inside A:C lane"),
        (15, 4, "roundRect", "Outside lane"),
    ]
    assert _drawing_state(output_path, "drawing2.xml") == [
        (26, 0, "rightArrow", "Inside A:C lane"),
        (15, 4, "roundRect", "Outside lane"),
    ]
    assert_no_template_tags(output_path)
    return output_path


def main() -> None:
    """Generate the template and rendered output workbooks."""

    template = build_template()
    output = render_sample(template)
    print(f"Generated {template.name} and {output.name}")


if __name__ == "__main__":
    main()
