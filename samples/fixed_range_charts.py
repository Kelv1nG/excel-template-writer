"""Generate the maintained sample for fixed-range template-authored charts."""

from __future__ import annotations

from pathlib import Path
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
from openpyxl.worksheet.worksheet import Worksheet

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    BLUE,
    LIGHT_BLUE,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("fixed_range_charts")


def _chart_references(chart: BarChart) -> set[str]:
    """Return every fixed formula serialized by one chart.

    Args:
        chart: Loaded chart to serialize.

    Returns:
        Set of direct worksheet formulas stored by the chart.
    """

    tree = chart._write()
    return {
        element.text
        for element in tree.iter()
        if element.tag.rsplit("}", 1)[-1] == "f" and element.text
    }


def _chart_style(path: Path, part: str) -> str | None:
    """Read the chart-space style that openpyxl omits from its loaded chart object.

    Args:
        path: Workbook containing the chart XML part.
        part: Chart XML filename inside ``xl/charts``.

    Returns:
        Serialized chart-style identifier, or ``None`` when absent.
    """

    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(f"xl/charts/{part}"))
    style = next(
        (element for element in root.iter() if element.tag.rsplit("}", 1)[-1] == "style"),
        None,
    )
    return None if style is None else style.get("val")


def _add_fixed_table(sheet: Worksheet, *, shift: str) -> None:
    """Add the twelve-row repeat source used by one chart example.

    Args:
        sheet: Worksheet receiving the table template.
        shift: Either the default whole-row policy or the cell-lane policy.
    """

    sheet["A4"] = "Identifier"
    sheet["B4"] = "Plan"
    sheet["C4"] = "Actual"
    paint(sheet, "A4:C4", fill=BLUE, bold=True, font_color=WHITE, horizontal="center")

    shift_option = "" if shift == "rows" else ' shift="cells"'
    sheet["A5"] = f"{{% for item in items{shift_option} %}}{{{{ item.identifier }}}}"
    sheet["B5"] = "{{ item.plan }}"
    sheet["C5"] = "{{ item.actual }}{% endfor %}"
    paint(sheet, "A5:C5", fill=LIGHT_BLUE)
    sheet["B5"].number_format = "$#,##0"
    sheet["C5"].number_format = "$#,##0"


def _add_fixed_chart(sheet: Worksheet, *, anchor: str, title: str) -> None:
    """Add one fixed-nine-row chart at the requested template anchor.

    Args:
        sheet: Worksheet receiving the template-authored chart.
        anchor: Top-left chart anchor in A1 notation.
        title: Visible chart title.
    """

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    cast(Any, chart.y_axis).title = "Amount"
    chart.height = 9
    chart.width = 18
    chart.display_blanks = "gap"
    cast(Any, chart.legend).position = "b"
    chart.add_data(
        Reference(sheet, min_col=2, max_col=3, min_row=4, max_row=13),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=5, max_row=13))
    sheet.add_chart(chart, anchor)


def _prepare_example_sheet(
    sheet: Worksheet,
    *,
    title: str,
    description: str,
    shift: str,
    anchor: str,
    chart_title: str,
) -> None:
    """Build one complete fixed-range chart example sheet.

    Args:
        sheet: Worksheet to prepare.
        title: Visible worksheet example title.
        description: Visible explanation of its anchor behavior.
        shift: Repeat shift policy, either ``rows`` or ``cells``.
        anchor: Authored chart anchor in A1 notation.
        chart_title: Visible title inside the chart.
    """

    prepare_sheet(
        sheet,
        title,
        description,
        widths=(22, 15, 15, 3, 14, 14, 14, 14, 14, 14, 14, 14),
    )
    _add_fixed_table(sheet, shift=shift)
    _add_fixed_chart(sheet, anchor=anchor, title=chart_title)


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build fixed-range charts with stationary and downward-moving anchors.

    Args:
        path: Destination for the generated template workbook.

    Returns:
        Verified template path.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fixed chart"
    _prepare_example_sheet(
        sheet,
        title="Fixed-range template chart",
        description=("A chart beside a cell-shift table stays put while its ranges remain fixed."),
        shift="cells",
        anchor="E4",
        chart_title="First nine items only",
    )

    rows_sheet = workbook.create_sheet("Pushed - rows")
    _prepare_example_sheet(
        rows_sheet,
        title="Chart pushed by row expansion",
        description=("The chart starts at E16 and moves to E27 when twelve rows replace one."),
        shift="rows",
        anchor="E16",
        chart_title="Moves with whole rows",
    )

    cells_sheet = workbook.create_sheet("Pushed - cells")
    _prepare_example_sheet(
        cells_sheet,
        title="Chart pushed inside a cell lane",
        description=("The chart starts at A16 and moves to A27 inside the expanding A:C lane."),
        shift="cells",
        anchor="A16",
        chart_title="Moves inside the table lane",
    )

    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render twelve rows and verify fixed ranges plus planned chart movement.

    Args:
        template_path: Authored source workbook.
        output_path: Separate rendered workbook destination.

    Returns:
        Verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {
            "items": [
                {
                    "identifier": f"Item {index:02d}",
                    "plan": 1_000 + index * 80,
                    "actual": 950 + index * 95,
                }
                for index in range(1, 13)
            ]
        },
    )

    workbook = load_workbook(output_path)
    try:
        for sheet_name, expected_anchor in (
            ("Fixed chart", (3, 4)),
            ("Pushed - rows", (26, 4)),
            ("Pushed - cells", (26, 0)),
        ):
            sheet = workbook[sheet_name]
            assert [sheet[f"A{row}"].value for row in range(5, 17)] == [
                f"Item {index:02d}" for index in range(1, 13)
            ]
            assert len(sheet._charts) == 1
            chart = sheet._charts[0]
            assert type(chart) is BarChart
            assert chart.display_blanks == "gap"
            assert chart.title is not None
            assert cast(Any, chart.y_axis).title is not None
            assert cast(Any, chart.legend).position == "b"
            assert isinstance(chart.anchor, OneCellAnchor)
            assert (chart.anchor._from.row, chart.anchor._from.col) == expected_anchor
            assert _chart_references(chart) == {
                f"'{sheet_name}'!B4",
                f"'{sheet_name}'!C4",
                f"'{sheet_name}'!$A$5:$A$13",
                f"'{sheet_name}'!$B$5:$B$13",
                f"'{sheet_name}'!$C$5:$C$13",
            }
        for part in ("chart1.xml", "chart2.xml", "chart3.xml"):
            assert _chart_style(template_path, part) == "10"
            assert _chart_style(output_path, part) == "10"
    finally:
        workbook.close()
    assert_no_template_tags(output_path)
    return output_path


def main() -> None:
    """Generate the template and rendered output workbooks."""

    template = build_template()
    output = render_sample(template)
    print(f"Generated {template.name} and {output.name}")


if __name__ == "__main__":
    main()
