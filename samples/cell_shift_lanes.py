"""Generate the maintained sample for independently growing cell-shift lanes."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    LIGHT_BLUE,
    LIGHT_GOLD,
    LIGHT_GRAY,
    LIGHT_GREEN,
    NAVY,
    SAMPLES_DIR,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("cell_shift_lanes")


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build a template with two independent cell-shift repeats.

    Args:
        path: Destination path for the authored template.

    Returns:
        The verified template path.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cell-shift lanes"
    prepare_sheet(
        sheet,
        "Independent cell-shift lanes",
        "The left and right repeats grow independently while the gray middle columns stay fixed.",
        widths=(24, 16, 22, 22, 24, 16),
    )
    headers = ("Left lane", "Marker", "Stationary", "Stationary", "Right lane", "Marker")
    for column, label in enumerate(headers, start=1):
        sheet.cell(4, column, label)
    paint(sheet, "A4:F4", fill=NAVY, bold=True, font_color=WHITE, horizontal="center")
    sheet["A5"] = '{% for item in left_items shift="cells" %}{{ item }}'
    sheet["B5"] = "LEFT{% endfor %}"
    sheet["C5"] = "STAYS AT C5"
    sheet["D5"] = "STAYS AT D5"
    sheet["E5"] = '{% for item in right_items shift="cells" %}{{ item }}'
    sheet["F5"] = "RIGHT{% endfor %}"
    paint(sheet, "A5:B5", fill=LIGHT_BLUE, bold=True)
    paint(sheet, "C5:D5", fill=LIGHT_GRAY, bold=True, horizontal="center")
    paint(sheet, "E5:F5", fill=LIGHT_GREEN, bold=True)
    sheet.row_dimensions[5].height = None
    sheet["A6"] = "LEFT FOOTER — expected at A8"
    sheet["B6"] = "moves with left"
    sheet["C6"] = "STAYS AT C6"
    sheet["D6"] = "STAYS AT D6"
    sheet["E6"] = "RIGHT FOOTER — expected at E7"
    sheet["F6"] = "moves with right"
    paint(sheet, "A6:B6", fill=LIGHT_GOLD, bold=True)
    paint(sheet, "C6:D6", fill=LIGHT_GRAY, bold=True, horizontal="center")
    paint(sheet, "E6:F6", fill=LIGHT_GOLD, bold=True)
    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render and verify independent cell-shift behavior.

    Args:
        template_path: Authored sample template path.
        output_path: Separate rendered workbook path.

    Returns:
        The verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {"left_items": ["Left A", "Left B", "Left C"], "right_items": ["Right 1", "Right 2"]},
    )
    assert_no_template_tags(output_path)
    workbook = load_workbook(output_path)
    try:
        sheet = workbook["Cell-shift lanes"]
        assert [sheet[f"A{row}"].value for row in range(5, 8)] == [
            "Left A",
            "Left B",
            "Left C",
        ]
        assert [sheet[f"E{row}"].value for row in range(5, 7)] == ["Right 1", "Right 2"]
        assert sheet["A8"].value.startswith("LEFT FOOTER")
        assert sheet["E7"].value.startswith("RIGHT FOOTER")
        assert sheet["C5"].value == "STAYS AT C5"
        assert sheet["C6"].value == "STAYS AT C6"
    finally:
        workbook.close()
    return output_path


def main() -> None:
    """Generate both cell-shift sample workbooks and print their paths."""

    template = build_template()
    output = render_sample(template)
    print(f"Template: {template.relative_to(SAMPLES_DIR.parent)}")
    print(f"Output:   {output.relative_to(SAMPLES_DIR.parent)}")


if __name__ == "__main__":
    main()
