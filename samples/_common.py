"""Shared workbook-building helpers for maintained samples."""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

SAMPLES_DIR = Path(__file__).resolve().parent

NAVY = "FF17365D"
BLUE = "FF5B9BD5"
LIGHT_BLUE = "FFDDEBF7"
LIGHT_GREEN = "FFE2F0D9"
LIGHT_GOLD = "FFFFF2CC"
LIGHT_GRAY = "FFE7E6E6"
WHITE = "FFFFFFFF"
INK = "FF203040"
THIN = Side(style="thin", color="FF8EA9C1")
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sample_paths(name: str) -> tuple[Path, Path]:
    """Return conventional template and output paths for one sample.

    Args:
        name: Stable lowercase sample stem.

    Returns:
        The ``(template_path, output_path)`` pair inside ``samples/``.
    """

    return SAMPLES_DIR / f"{name}_template.xlsx", SAMPLES_DIR / f"{name}_output.xlsx"


def atomic_save(workbook: WorkbookType, path: Path) -> Path:
    """Save and reopen a generated workbook before replacing its destination.

    Args:
        workbook: In-memory workbook to serialize.
        path: Final workbook path.

    Returns:
        The verified destination path.
    """

    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(suffix=".xlsx", dir=path.parent)
    os.close(handle)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        load_workbook(temporary_path, read_only=True, data_only=False).close()
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)
    return path


def paint(
    sheet: Worksheet,
    cell_range: str,
    *,
    fill: str,
    bold: bool = False,
    font_color: str = INK,
    horizontal: str = "left",
    wrap: bool = True,
) -> None:
    """Apply the samples' standard direct cell presentation to a range.

    Args:
        sheet: Worksheet containing the range.
        cell_range: Excel A1 range to format.
        fill: Eight-digit ARGB fill color.
        bold: Whether text should be bold.
        font_color: Eight-digit ARGB font color.
        horizontal: Horizontal alignment name.
        wrap: Whether cell text should wrap.
    """

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Aptos", size=11, bold=bold, color=font_color)
            cell.border = GRID_BORDER
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def merge_band(
    sheet: Worksheet,
    cell_range: str,
    value: str,
    *,
    fill: str,
    bold: bool = False,
    font_color: str = INK,
    horizontal: str = "left",
) -> None:
    """Create and style one merged label band.

    Args:
        sheet: Worksheet receiving the merged range.
        cell_range: Excel A1 range to merge.
        value: Label written to the merged range's top-left cell.
        fill: Eight-digit ARGB fill color.
        bold: Whether text should be bold.
        font_color: Eight-digit ARGB font color.
        horizontal: Horizontal alignment name.
    """

    min_col, min_row, _, _ = range_boundaries(cell_range)
    sheet.merge_cells(cell_range)
    sheet.cell(min_row, min_col, value)
    paint(
        sheet,
        cell_range,
        fill=fill,
        bold=bold,
        font_color=font_color,
        horizontal=horizontal,
    )


def prepare_sheet(
    sheet: Worksheet,
    title: str,
    description: str,
    *,
    widths: tuple[float, ...],
) -> None:
    """Apply the common title, description, widths, and view settings.

    Args:
        sheet: Worksheet to initialize.
        title: Prominent sample title.
        description: Short explanation of the demonstrated behavior.
        widths: Column widths beginning at column A.
    """

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    last_column = get_column_letter(len(widths))
    merge_band(
        sheet,
        f"A1:{last_column}1",
        title,
        fill=NAVY,
        bold=True,
        font_color=WHITE,
        horizontal="center",
    )
    merge_band(
        sheet,
        f"A2:{last_column}2",
        description,
        fill=LIGHT_BLUE,
        horizontal="center",
    )
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 34


def assert_no_template_tags(path: Path) -> None:
    """Assert that a rendered workbook contains no template delimiters.

    Args:
        path: Rendered workbook path to reopen and inspect.
    """

    workbook = load_workbook(path, data_only=False)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        assert "{%" not in cell.value
                        assert "{{" not in cell.value
    finally:
        workbook.close()
