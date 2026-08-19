"""Generate maintained samples for explicit vertical layout regions."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    LIGHT_BLUE,
    LIGHT_GOLD,
    LIGHT_GREEN,
    NAVY,
    SAMPLES_DIR,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    merge_band,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("regions")


def _add_lane_headers(sheet: Worksheet) -> None:
    """Add shared lane labels to a region sample worksheet.

    Args:
        sheet: Openpyxl worksheet receiving lane labels.
    """

    merge_band(sheet, "A4:C4", "LEFT — 3 items", fill=LIGHT_BLUE, bold=True, horizontal="center")
    merge_band(sheet, "D4:F4", "MIDDLE — 5 items", fill=LIGHT_GREEN, bold=True, horizontal="center")
    merge_band(sheet, "G4:I4", "RIGHT — 2 items", fill=LIGHT_BLUE, bold=True, horizontal="center")
    merge_band(sheet, "K4:P4", "ADJACENT SECTION", fill=LIGHT_GOLD, bold=True, horizontal="center")


def _add_parallel_region(sheet: Worksheet, *, shift: str) -> None:
    """Add one A:J region containing three independent child lanes.

    Args:
        sheet: Openpyxl worksheet receiving the region.
        shift: External region shift policy, either ``rows`` or ``cells``.
    """

    sheet["A5"] = f'{{% region shift="{shift}" %}}A:J REGION'
    paint(sheet, "A5:J5", fill=NAVY, bold=True, font_color=WHITE)
    sheet["A6"] = '{% for item in left_items shift="cells" %}{{ item }}'
    sheet["C6"] = "LEFT{% endfor %}"
    sheet["D6"] = '{% for item in middle_items shift="cells" %}{{ item }}'
    sheet["F6"] = "MIDDLE{% endfor %}"
    sheet["G6"] = '{% for item in right_items shift="cells" %}{{ item }}'
    sheet["I6"] = "RIGHT{% endfor %}"
    sheet["J6"] = "{% endregion %}"
    paint(sheet, "A6:C6", fill=LIGHT_BLUE, bold=True)
    paint(sheet, "D6:F6", fill=LIGHT_GREEN, bold=True)
    paint(sheet, "G6:J6", fill=LIGHT_BLUE, bold=True)
    sheet.row_dimensions[6].height = None


def _build_cell_band_sheet(workbook: Workbook) -> None:
    """Add a cell-shift region whose adjacent columns stay fixed.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Region - cells")
    prepare_sheet(
        sheet,
        'Region shift="cells"',
        "The tallest child grows the full A:J band; K:P remains at its authored rows.",
        widths=(14,) * 16,
    )
    _add_lane_headers(sheet)
    _add_parallel_region(sheet, shift="cells")
    merge_band(
        sheet,
        "A10:J10",
        "A:J FOOTER — expected at row 14",
        fill=LIGHT_BLUE,
        bold=True,
        horizontal="center",
    )
    merge_band(
        sheet,
        "K10:P10",
        "K:P FOOTER — stays on row 10",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def _build_row_sheet(workbook: Workbook) -> None:
    """Add a row-shift region whose growth moves every column.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Region - rows")
    prepare_sheet(
        sheet,
        'Region shift="rows"',
        "The same completed region inserts complete rows, so both footer bands move together.",
        widths=(14,) * 16,
    )
    _add_lane_headers(sheet)
    _add_parallel_region(sheet, shift="rows")
    merge_band(
        sheet,
        "A10:J10",
        "A:J FOOTER — expected at row 14",
        fill=LIGHT_BLUE,
        bold=True,
        horizontal="center",
    )
    merge_band(
        sheet,
        "K10:P10",
        "K:P FOOTER — also expected at row 14",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def _build_reserved_height_sheet(workbook: Workbook) -> None:
    """Add a region whose source height absorbs child growth.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Region - reserved height")
    prepare_sheet(
        sheet,
        "Reserved region height",
        "A4:J12 reserves nine rows; five list items fit without moving the footer below row 15.",
        widths=(18,) * 10,
    )
    sheet["A4"] = '{% region shift="cells" %}RESERVED A4:J12'
    paint(sheet, "A4:J4", fill=NAVY, bold=True, font_color=WHITE)
    sheet["A5"] = '{% for item in middle_items shift="cells" %}{{ item }}{% endfor %}'
    paint(sheet, "A5:C5", fill=LIGHT_GREEN, bold=True)
    sheet["J12"] = "{% endregion %}"
    paint(sheet, "J12", fill=LIGHT_BLUE)
    merge_band(
        sheet,
        "A15:J15",
        "FOOTER — remains on row 15",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def _build_nested_sheet(workbook: Workbook) -> None:
    """Add nested regions that measure from the inside out.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Region - nested")
    prepare_sheet(
        sheet,
        "Nested regions",
        "The inner A5:F6 region completes before the outer A4:J6 region moves its A:J footer.",
        widths=(16,) * 12,
    )
    sheet["A4"] = '{% region shift="cells" %}OUTER REGION'
    paint(sheet, "A4:J4", fill=NAVY, bold=True, font_color=WHITE)
    sheet["A5"] = '{% region shift="cells" %}INNER REGION'
    paint(sheet, "A5:F5", fill=LIGHT_BLUE, bold=True)
    sheet["A6"] = '{% for item in middle_items shift="cells" %}{{ item }}{% endfor %}'
    sheet["F6"] = "{% endregion %}"
    sheet["J6"] = "{% endregion %}"
    paint(sheet, "A6:F6", fill=LIGHT_GREEN)
    merge_band(
        sheet,
        "A10:J10",
        "OUTER FOOTER — expected at row 14",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    sheet["K10"] = "Adjacent cell stays on row 10"
    paint(sheet, "K10:L10", fill=LIGHT_BLUE)


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build the complete explicit-regions template workbook.

    Args:
        path: Destination path for the authored template.

    Returns:
        The verified template path.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_cell_band_sheet(workbook)
    _build_row_sheet(workbook)
    _build_reserved_height_sheet(workbook)
    _build_nested_sheet(workbook)
    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render and verify every explicit-region behavior.

    Args:
        template_path: Authored sample template path.
        output_path: Separate rendered workbook path.

    Returns:
        The verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {
            "left_items": ["Left A", "Left B", "Left C"],
            "middle_items": ["Middle 1", "Middle 2", "Middle 3", "Middle 4", "Middle 5"],
            "right_items": ["Right 1", "Right 2"],
        },
    )
    assert_no_template_tags(output_path)
    workbook = load_workbook(output_path)
    try:
        cells = workbook["Region - cells"]
        assert [cells[f"D{row}"].value for row in range(6, 11)] == [
            "Middle 1",
            "Middle 2",
            "Middle 3",
            "Middle 4",
            "Middle 5",
        ]
        assert cells["A14"].value.startswith("A:J FOOTER")
        assert cells["K10"].value.startswith("K:P FOOTER")
        rows = workbook["Region - rows"]
        assert rows["A14"].value.startswith("A:J FOOTER")
        assert rows["K14"].value.startswith("K:P FOOTER")
        reserved = workbook["Region - reserved height"]
        assert reserved["A15"].value.startswith("FOOTER")
        nested = workbook["Region - nested"]
        assert nested["A14"].value.startswith("OUTER FOOTER")
        assert nested["K10"].value == "Adjacent cell stays on row 10"
    finally:
        workbook.close()
    return output_path


def main() -> None:
    """Generate both explicit-region sample workbooks and print their paths."""

    template = build_template()
    output = render_sample(template)
    print(f"Template: {template.relative_to(SAMPLES_DIR.parent)}")
    print(f"Output:   {output.relative_to(SAMPLES_DIR.parent)}")


if __name__ == "__main__":
    main()
