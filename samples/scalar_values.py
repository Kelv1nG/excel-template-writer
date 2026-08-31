"""Generate the maintained scalar values and expressions sample workbooks."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    LIGHT_BLUE,
    LIGHT_GOLD,
    SAMPLES_DIR,
    assert_no_template_tags,
    atomic_save,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("scalar_values")


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build a template demonstrating typed cells, mixed text, and scalar filters.

    Args:
        path: Destination path for the authored template workbook.

    Returns:
        The verified template path.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scalar values"
    prepare_sheet(
        sheet,
        "Scalar values and expressions",
        "Typed expressions support explicit filters, collection aggregates, and basic arithmetic.",
        widths=(24, 62, 34),
    )
    sheet.append(["Case", "Template value", "Expected behavior"])
    paint(sheet, "A4:C4", fill=LIGHT_BLUE, bold=True, horizontal="center")
    rows = (
        ("Text", "{{ customer.name }}", "native text"),
        ("Integer", "{{ invoice.quantity }}", "native number"),
        ("Decimal", "{{ invoice.amount }}", "currency-formatted number"),
        ("Date", "{{ invoice.issued_on }}", "native date"),
        (
            "Date text filter",
            '{{ invoice.issued_on | date("YYYY-mm") }}',
            "explicit text: 2026-08",
        ),
        (
            "Date in sentence",
            'For the month ending {{ invoice.issued_on | date("dd mmmm yyyy") }}',
            "formatted text in mixed content",
        ),
        ("Boolean", "{{ invoice.approved }}", "native boolean"),
        ("Mixed text", "Invoice {{ invoice.number }}", "always text"),
        ("Upper filter", "{{ customer.name | upper }}", "uppercase text"),
        ("Join filter", '{{ labels | join(" / ") }}', "one text value"),
        ("Numeric sum", "{{ amounts | sum }}", "nulls skipped; native number: 25"),
        (
            "Record-column sum",
            '{{ lines | sum("amount") }}',
            "literal column key; native number: 150",
        ),
        ("Record-column minimum", '{{ lines | min("amount") }}', "nulls skipped: 50"),
        ("Record-column maximum", '{{ lines | max("amount") }}', "nulls skipped: 100"),
        ("Record count", "{{ lines | count }}", "all records counted: 3"),
        (
            "Aggregate arithmetic",
            '{{ (lines | max("amount")) - (lines | min("amount")) }}',
            "parenthesized filtered operands: 50",
        ),
        (
            "Basic arithmetic",
            "{{ ((math.base + math.adjustment) * math.factor - math.discount) / math.divisor }}",
            "standard precedence: 50",
        ),
        ("Unary signs", "{{ -math.adjustment + +math.base }}", "native number: 80"),
        ("Default filter", '{{ missing | default("not supplied") }}', "explicit fallback"),
    )
    for row_number, values in enumerate(rows, start=5):
        for column, value in enumerate(values, start=1):
            sheet.cell(row_number, column, value)
        paint(
            sheet, f"A{row_number}:C{row_number}", fill=LIGHT_GOLD if row_number % 2 else "FFFFFFFF"
        )
    sheet["B7"].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    sheet["B8"].number_format = "yyyy-mm-dd"
    sheet["B16"].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    for coordinate in ("B17", "B18", "B20", "B21"):
        sheet[coordinate].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render and verify the scalar sample workbook.

    Args:
        template_path: Authored sample template path.
        output_path: Separate rendered workbook path.

    Returns:
        The verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {
            "customer": {"name": "Acme Industries"},
            "invoice": {
                "number": "INV-1042",
                "quantity": 3,
                "amount": 1250.75,
                "issued_on": date(2026, 8, 19),
                "approved": True,
            },
            "labels": ["priority", "renewal"],
            "amounts": [10, None, 15],
            "lines": [{"amount": 100}, {"amount": None}, {"amount": 50}],
            "math": {
                "base": 100,
                "adjustment": 20,
                "factor": 2,
                "discount": 40,
                "divisor": 4,
            },
        },
    )
    assert_no_template_tags(output_path)
    workbook = load_workbook(output_path)
    try:
        sheet = workbook["Scalar values"]
        assert sheet["B5"].value == "Acme Industries"
        assert sheet["B6"].value == 3
        assert sheet["B7"].value == 1250.75
        assert sheet["B8"].data_type == "d"
        assert sheet["B9"].value == "2026-08"
        assert sheet["B9"].data_type == "s"
        assert sheet["B10"].value == "For the month ending 19 August 2026"
        assert sheet["B10"].data_type == "s"
        assert sheet["B11"].value is True
        assert sheet["B12"].value == "Invoice INV-1042"
        assert sheet["B15"].value == 25
        assert sheet["B15"].data_type == "n"
        assert sheet["B16"].value == 150
        assert sheet["B16"].data_type == "n"
        assert sheet["B16"].number_format == '$#,##0.00;[Red]-$#,##0.00;"-"'
        assert sheet["B17"].value == 50
        assert sheet["B17"].data_type == "n"
        assert sheet["B18"].value == 100
        assert sheet["B18"].data_type == "n"
        assert sheet["B19"].value == 3
        assert sheet["B19"].data_type == "n"
        assert sheet["B20"].value == 50
        assert sheet["B20"].data_type == "n"
        assert sheet["B21"].value == 50
        assert sheet["B21"].data_type == "n"
        assert sheet["B22"].value == 80
        assert sheet["B22"].data_type == "n"
        assert sheet["B23"].value == "not supplied"
    finally:
        workbook.close()
    return output_path


def main() -> None:
    """Generate both scalar sample workbooks and print their paths."""

    template = build_template()
    output = render_sample(template)
    print(f"Template: {template.relative_to(SAMPLES_DIR.parent)}")
    print(f"Output:   {output.relative_to(SAMPLES_DIR.parent)}")


if __name__ == "__main__":
    main()
