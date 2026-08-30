"""Generate maintained examples for template-authored embedded image movement."""

from __future__ import annotations

import struct
import zlib
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    BLUE,
    LIGHT_BLUE,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("template_images")
_DRAWING_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"


def _png_chunk(chunk_type: bytes, data: bytes) -> bytes:
    """Build one checksummed PNG chunk.

    Args:
        chunk_type: Four-byte PNG chunk identifier.
        data: Raw chunk payload.

    Returns:
        Complete length-prefixed and checksummed chunk.
    """

    checksum = zlib.crc32(chunk_type)
    checksum = zlib.crc32(data, checksum)
    return struct.pack(">I", len(data)) + chunk_type + data + struct.pack(">I", checksum)


def _solid_png(width: int, height: int, color: tuple[int, int, int]) -> bytes:
    """Create a small solid-color RGB PNG without an image-library dependency.

    Args:
        width: Image width in pixels.
        height: Image height in pixels.
        color: Red, green, and blue channel values.

    Returns:
        Valid PNG media bytes.
    """

    scanline = b"\x00" + bytes(color) * width
    pixels = scanline * height
    header = struct.pack(">IIBBBBB", width, height, 8, 2, 0, 0, 0)
    return b"\x89PNG\r\n\x1a\n" + b"".join(
        (
            _png_chunk(b"IHDR", header),
            _png_chunk(b"IDAT", zlib.compress(pixels)),
            _png_chunk(b"IEND", b""),
        )
    )


class _EmbeddedPng(Image):
    def __init__(self, data: bytes, *, width: int = 180, height: int = 80) -> None:
        """Create an openpyxl-compatible in-memory PNG without Pillow.

        Args:
            data: Complete PNG payload.
            width: Display width in pixels.
            height: Display height in pixels.
        """

        self._payload = data
        self.width = width
        self.height = height
        self.format = "png"

    def _data(self) -> bytes:
        """Return the embedded PNG payload."""

        return self._payload


BLUE_IMAGE = _solid_png(180, 80, (91, 155, 213))
GREEN_IMAGE = _solid_png(180, 80, (112, 173, 71))


def _add_table(sheet: Worksheet, *, shift: str) -> None:
    """Add the one-row table source used by an image movement sheet.

    Args:
        sheet: Worksheet receiving the table template.
        shift: Either the default ``rows`` policy or isolated ``cells`` policy.
    """

    sheet["A4"] = "Identifier"
    sheet["B4"] = "Plan"
    sheet["C4"] = "Actual"
    paint(sheet, "A4:C4", fill=BLUE, bold=True, font_color=WHITE, horizontal="center")
    shift_option = "" if shift == "rows" else ' shift="cells"'
    sheet["A5"] = f"{{% for item in items{shift_option} %}}{{{{ item.identifier }}}}"
    sheet["B5"] = "{{ item.plan }}"
    sheet["C5"] = "{{ item.actual }}{% endfor %}"
    paint(sheet, "A5:C5", fill=LIGHT_BLUE)
    sheet["B5"].number_format = "$#,##0"
    sheet["C5"].number_format = "$#,##0"


def _prepare_sheet(sheet: Worksheet, *, shift: str, description: str) -> None:
    """Prepare one visible image-anchor movement example.

    Args:
        sheet: Worksheet to initialize.
        shift: Table shift policy for this example.
        description: Visible explanation of the expected movement.
    """

    prepare_sheet(
        sheet,
        "Template-authored image movement",
        description,
        widths=(22, 15, 15, 3, 22, 14, 14, 14),
    )
    _add_table(sheet, shift=shift)


def _drawing_markers(path: Path, part: str) -> list[tuple[int, int]]:
    """Read zero-based one-cell image markers from one drawing part.

    Args:
        path: XLSX package containing the drawing.
        part: Drawing XML filename inside ``xl/drawings``.

    Returns:
        Image marker rows and columns in serialized stacking order.
    """

    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(f"xl/drawings/{part}"))
    markers: list[tuple[int, int]] = []
    for anchor in root:
        marker = anchor.find(f"{{{_DRAWING_NAMESPACE}}}from")
        assert marker is not None
        row = marker.find(f"{{{_DRAWING_NAMESPACE}}}row")
        column = marker.find(f"{{{_DRAWING_NAMESPACE}}}col")
        assert row is not None and row.text is not None
        assert column is not None and column.text is not None
        markers.append((int(row.text), int(column.text)))
    return markers


def _media_payloads(path: Path) -> list[bytes]:
    """Read all embedded media payloads in stable part-name order.

    Args:
        path: XLSX package containing embedded images.

    Returns:
        Raw media bytes sorted by their package path.
    """

    with ZipFile(path) as archive:
        names = sorted(name for name in archive.namelist() if name.startswith("xl/media/"))
        return [archive.read(name) for name in names]


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build image templates for whole-row and cell-lane movement.

    Args:
        path: Destination for the generated template workbook.

    Returns:
        Verified template path.
    """

    workbook = Workbook()
    row_sheet = workbook.active
    row_sheet.title = "Row shift"
    _prepare_sheet(
        row_sheet,
        shift="rows",
        description="The E4 image stays fixed; the image at E16 moves to E27.",
    )
    row_sheet.add_image(_EmbeddedPng(BLUE_IMAGE), "E4")
    row_sheet.add_image(_EmbeddedPng(GREEN_IMAGE), "E16")

    lane_sheet = workbook.create_sheet("Cell lane")
    _prepare_sheet(
        lane_sheet,
        shift="cells",
        description="The A16 image moves to A27; the E16 image remains outside the A:C lane.",
    )
    lane_sheet.add_image(_EmbeddedPng(GREEN_IMAGE), "A16")
    lane_sheet.add_image(_EmbeddedPng(BLUE_IMAGE), "E16")

    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render twelve table rows and verify image movement plus media identity.

    Args:
        template_path: Authored source workbook.
        output_path: Separate rendered workbook destination.

    Returns:
        Verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {
            "items": [
                {
                    "identifier": f"Item {index:02d}",
                    "plan": 1_000 + index * 80,
                    "actual": 950 + index * 95,
                }
                for index in range(1, 13)
            ]
        },
    )

    workbook = load_workbook(output_path, read_only=True, data_only=False)
    try:
        for sheet_name in ("Row shift", "Cell lane"):
            sheet = workbook[sheet_name]
            assert [sheet[f"A{row}"].value for row in range(5, 17)] == [
                f"Item {index:02d}" for index in range(1, 13)
            ]
    finally:
        workbook.close()

    assert _drawing_markers(template_path, "drawing1.xml") == [(3, 4), (15, 4)]
    assert _drawing_markers(output_path, "drawing1.xml") == [(3, 4), (26, 4)]
    assert _drawing_markers(template_path, "drawing2.xml") == [(15, 0), (15, 4)]
    assert _drawing_markers(output_path, "drawing2.xml") == [(26, 0), (15, 4)]
    assert _media_payloads(output_path) == _media_payloads(template_path)
    assert_no_template_tags(output_path)
    return output_path


def main() -> None:
    """Generate the template and rendered output workbooks."""

    template = build_template()
    output = render_sample(template)
    print(f"Generated {template.name} and {output.name}")


if __name__ == "__main__":
    main()
