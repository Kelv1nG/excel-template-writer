"""Generate maintained samples for conditional and nested structural blocks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    LIGHT_BLUE,
    LIGHT_GOLD,
    LIGHT_GRAY,
    LIGHT_GREEN,
    NAVY,
    SAMPLES_DIR,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    merge_band,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("conditions_and_nesting")


def _build_conditions_sheet(workbook: Workbook) -> None:
    """Add stacked true/false branch examples.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Conditions")
    prepare_sheet(
        sheet,
        "Rectangular conditions",
        "Only the selected branch remains, and the unselected branch's height is removed.",
        widths=(28, 18, 18, 18),
    )
    sheet["A4"] = "{% if account.active and not account.suspended %}Account"
    sheet["D4"] = "ACTIVE{% else %}"
    sheet["A5"] = "Account"
    sheet["D5"] = "INACTIVE{% endif %}"
    paint(sheet, "A4:D4", fill=LIGHT_BLUE, bold=True)
    paint(sheet, "A5:D5", fill=LIGHT_GRAY)
    sheet["A6"] = "{% if account.overdue %}Payment overdue{% endif %}"
    paint(sheet, "A6:D6", fill=LIGHT_GOLD, bold=True)
    merge_band(
        sheet,
        "A7:D7",
        "Footer compacts after branch selection",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def _build_nesting_sheet(workbook: Workbook) -> None:
    """Add nested group and item repeats with conditional content.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Nested blocks")
    prepare_sheet(
        sheet,
        "Nested repeats and lexical scope",
        "Each group measures its own item repeat before the outer instances are stacked.",
        widths=(32, 22, 18, 18),
    )
    sheet.merge_cells("A4:D4")
    sheet["A4"] = "{% for group in groups %}{{ group.name }}"
    paint(sheet, "A4:D4", fill=NAVY, bold=True, font_color=WHITE)
    sheet["A5"] = "{% for item in group.items %}{{ item.name }}"
    sheet["B5"] = "{{ item.kind }}"
    sheet["D5"] = "ITEM{% endfor %}"
    paint(sheet, "A5:D5", fill=LIGHT_BLUE)
    sheet["A6"] = "Group count"
    sheet["D6"] = "{{ group.count }}{% endfor %}"
    paint(sheet, "A6:D6", fill=LIGHT_GREEN, bold=True)
    merge_band(
        sheet,
        "A7:D7",
        "Footer moves below both fully measured groups",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build the conditions-and-nesting template workbook.

    Args:
        path: Destination path for the authored template.

    Returns:
        The verified template path.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_conditions_sheet(workbook)
    _build_nesting_sheet(workbook)
    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render and verify the condition and nesting examples.

    Args:
        template_path: Authored sample template path.
        output_path: Separate rendered workbook path.

    Returns:
        The verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {
            "account": {"active": True, "suspended": False, "overdue": False},
            "groups": [
                {
                    "name": "Hardware",
                    "count": 2,
                    "items": [
                        {"name": "Laptop", "kind": "Highlighted"},
                        {"name": "Monitor", "kind": "Standard"},
                    ],
                },
                {
                    "name": "Software",
                    "count": 3,
                    "items": [
                        {"name": "Editor", "kind": "Standard"},
                        {"name": "Database", "kind": "Highlighted"},
                        {"name": "Support", "kind": "Standard"},
                    ],
                },
            ],
        },
    )
    assert_no_template_tags(output_path)
    workbook = load_workbook(output_path)
    try:
        conditions = workbook["Conditions"]
        assert conditions["D4"].value == "ACTIVE"
        assert conditions["A5"].value.startswith("Footer compacts")
        nested = workbook["Nested blocks"]
        assert nested["A4"].value == "Hardware"
        assert nested["A5"].value == "Laptop"
        assert nested["B5"].value == "Highlighted"
        assert nested["A8"].value == "Software"
        assert nested["A12"].value == "Group count"
        assert nested["A13"].value.startswith("Footer moves")
    finally:
        workbook.close()
    return output_path


def main() -> None:
    """Generate both condition/nesting sample workbooks and print their paths."""

    template = build_template()
    output = render_sample(template)
    print(f"Template: {template.relative_to(SAMPLES_DIR.parent)}")
    print(f"Output:   {output.relative_to(SAMPLES_DIR.parent)}")


if __name__ == "__main__":
    main()
