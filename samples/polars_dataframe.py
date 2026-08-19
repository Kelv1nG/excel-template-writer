"""Generate the optional maintained sample for eager Polars data frames."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import polars as pl
from openpyxl import Workbook, load_workbook

from excel_template_writer.adapters.polars import polars_adapters
from excel_template_writer.xlsx import render_workbook
from samples._common import (
    LIGHT_BLUE,
    LIGHT_GOLD,
    NAVY,
    SAMPLES_DIR,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    merge_band,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("polars_dataframe")


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build a styled table template consumed by a Polars adapter.

    Args:
        path: Destination path for the authored template.

    Returns:
        The verified template path.
    """

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Polars DataFrame"
    prepare_sheet(
        sheet,
        "Eager Polars DataFrame adapter",
        "The adapter preserves row order and column names; template syntax remains "
        "library-neutral.",
        widths=(30, 18, 18, 18),
    )
    for column, label in enumerate(("Description", "Amount", "Service date", "Optional"), start=1):
        sheet.cell(4, column, label)
    paint(sheet, "A4:D4", fill=NAVY, bold=True, font_color=WHITE, horizontal="center")
    sheet["A5"] = "{% for line in lines %}{{ line.description }}"
    sheet["B5"] = "{{ line.amount }}"
    sheet["C5"] = "{{ line.service_date }}"
    sheet["D5"] = "{{ line.optional }}{% endfor %}"
    paint(sheet, "A5:D5", fill=LIGHT_BLUE)
    sheet["B5"].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    sheet["C5"].number_format = "yyyy-mm-dd"
    merge_band(
        sheet,
        "A6:D6",
        "Footer moves below DataFrame rows",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render and verify an eager Polars DataFrame through explicit adapters.

    Args:
        template_path: Authored sample template path.
        output_path: Separate rendered workbook path.

    Returns:
        The verified rendered workbook path.
    """

    frame = pl.DataFrame(
        {
            "description": ["Consulting", "Support", "Renewal"],
            "amount": [1250.0, 350.0, 875.0],
            "service_date": [date(2026, 8, 19), date(2026, 8, 20), date(2026, 8, 21)],
            "optional": [1.0, None, float("nan")],
        },
        schema_overrides={"service_date": pl.Date},
    )
    render_workbook(
        template_path,
        output_path,
        {"lines": frame},
        adapters=polars_adapters(),
    )
    assert_no_template_tags(output_path)
    workbook = load_workbook(output_path)
    try:
        sheet = workbook["Polars DataFrame"]
        assert [sheet[f"A{row}"].value for row in range(5, 8)] == [
            "Consulting",
            "Support",
            "Renewal",
        ]
        assert sheet["B7"].value == 875
        assert sheet["C5"].data_type == "d"
        assert sheet["D6"].value is None
        assert sheet["D7"].value is None
        assert sheet["A8"].value.startswith("Footer moves")
    finally:
        workbook.close()
    return output_path


def main() -> None:
    """Generate both Polars sample workbooks and print their paths."""

    template = build_template()
    output = render_sample(template)
    print(f"Template: {template.relative_to(SAMPLES_DIR.parent)}")
    print(f"Output:   {output.relative_to(SAMPLES_DIR.parent)}")


if __name__ == "__main__":
    main()
