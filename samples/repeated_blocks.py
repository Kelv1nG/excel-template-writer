"""Generate maintained samples for lists, table rows, cards, and empty repeats."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from excel_template_writer.xlsx import render_workbook
from samples._common import (
    LIGHT_BLUE,
    LIGHT_GOLD,
    LIGHT_GREEN,
    NAVY,
    SAMPLES_DIR,
    WHITE,
    assert_no_template_tags,
    atomic_save,
    merge_band,
    paint,
    prepare_sheet,
    sample_paths,
)

TEMPLATE_PATH, OUTPUT_PATH = sample_paths("repeated_blocks")


def _build_table_sheet(workbook: Workbook) -> None:
    """Add a styled one-row table-body repeat.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Styled table")
    prepare_sheet(
        sheet,
        "Styled rectangular table body",
        "One authored row becomes three rows; values, styles, blank cells, and the footer move.",
        widths=(30, 12, 18, 16),
    )
    for column, label in enumerate(("Description", "Quantity", "Unit price", "Approved"), start=1):
        sheet.cell(4, column, label)
    paint(sheet, "A4:D4", fill=NAVY, bold=True, font_color=WHITE, horizontal="center")
    sheet["A5"] = "{% for line in lines %}{{ line.description }}"
    sheet["B5"] = "{{ line.quantity }}"
    sheet["C5"] = "{{ line.unit_price }}"
    sheet["D5"] = "{{ line.approved }}{% endfor %}"
    paint(sheet, "A5:D5", fill=LIGHT_BLUE)
    sheet["A5"].font = Font(name="Aptos", bold=True, color="FF17365D")
    sheet["C5"].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    merge_band(
        sheet,
        "A6:D6",
        "Footer moves below the completed table body",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def _build_list_sheet(workbook: Workbook) -> None:
    """Add a one-cell vertical list repeat.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("One-cell list")
    prepare_sheet(
        sheet,
        "Lists use the same repeat primitive",
        "A one-cell rectangular block renders an ordered list vertically.",
        widths=(30, 30),
    )
    sheet["A4"] = "{% for tag in tags %}{{ tag }}{% endfor %}"
    paint(sheet, "A4", fill=LIGHT_GREEN, bold=True)
    sheet["B4"] = "Static neighbor in the first source row"
    paint(sheet, "B4", fill=LIGHT_GOLD)
    merge_band(
        sheet,
        "A5:B5",
        "Footer moves to row 7 because the repeat shifts complete rows",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def _build_empty_sheet(workbook: Workbook) -> None:
    """Add an empty-repeat formatted placeholder example.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Empty repeat")
    prepare_sheet(
        sheet,
        "Empty collection placeholder",
        "An empty collection retains one formatted source instance; dynamic values become blank.",
        widths=(30, 28, 30),
    )
    sheet["A4"] = "{% for row in empty_rows %}{{ row.name }}"
    sheet["B4"] = None
    sheet["C4"] = "STATIC PLACEHOLDER{% endfor %}"
    paint(sheet, "A4:C4", fill=LIGHT_GREEN)
    sheet["B4"].fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    merge_band(
        sheet,
        "A5:C5",
        "Footer stays on row 5 because one source-sized instance remains",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def _build_directive_style_sheet(workbook: Workbook) -> None:
    """Show that directive-only cells retain presentation after tag removal.

    Args:
        workbook: Workbook receiving the worksheet.
    """

    sheet = workbook.create_sheet("Directive-only styles")
    prepare_sheet(
        sheet,
        "Directive-only cells remain formatted blanks",
        "The opening and closing tag cells keep their authored white fill and borders.",
        widths=(18, 32, 18),
    )
    sheet["A4"] = "{% for item in style_rows %}"
    sheet["B4"] = "{{ item }}"
    sheet["C4"] = "{% endfor %}"
    paint(sheet, "A4:C4", fill=WHITE, horizontal="center")
    merge_band(
        sheet,
        "A5:C5",
        "Footer moves below all four fully formatted rows",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )


def build_template(path: Path = TEMPLATE_PATH) -> Path:
    """Build the complete repeated-blocks template workbook.

    Args:
        path: Destination path for the authored template.

    Returns:
        The verified template path.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_table_sheet(workbook)
    _build_list_sheet(workbook)
    _build_empty_sheet(workbook)
    _build_directive_style_sheet(workbook)
    result = atomic_save(workbook, path)
    workbook.close()
    return result


def render_sample(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """Render and verify all repeated-block examples.

    Args:
        template_path: Authored sample template path.
        output_path: Separate rendered workbook path.

    Returns:
        The verified rendered workbook path.
    """

    render_workbook(
        template_path,
        output_path,
        {
            "lines": [
                {
                    "description": "Consulting",
                    "quantity": 2,
                    "unit_price": 1500.0,
                    "approved": True,
                },
                {
                    "description": "Implementation",
                    "quantity": 1,
                    "unit_price": 4250.0,
                    "approved": False,
                },
                {"description": "Support", "quantity": 6, "unit_price": 350.0, "approved": True},
            ],
            "tags": ["priority", "renewal", "enterprise"],
            "empty_rows": [],
            "style_rows": ["One", "Two", "Three", "Four"],
        },
    )
    assert_no_template_tags(output_path)
    workbook = load_workbook(output_path)
    try:
        table = workbook["Styled table"]
        assert [table[f"A{row}"].value for row in range(5, 8)] == [
            "Consulting",
            "Implementation",
            "Support",
        ]
        assert table["A8"].value.startswith("Footer moves")
        assert table["C7"].number_format == '$#,##0.00;[Red]-$#,##0.00;"-"'
        listing = workbook["One-cell list"]
        assert [listing[f"A{row}"].value for row in range(4, 7)] == [
            "priority",
            "renewal",
            "enterprise",
        ]
        assert listing["A7"].value.startswith("Footer moves")
        empty = workbook["Empty repeat"]
        assert empty["A4"].value is None
        assert empty["B4"].fill.fgColor.rgb == LIGHT_GOLD
        assert empty["C4"].value == "STATIC PLACEHOLDER"
        directive_styles = workbook["Directive-only styles"]
        assert [directive_styles[f"B{row}"].value for row in range(4, 8)] == [
            "One",
            "Two",
            "Three",
            "Four",
        ]
        for row in range(4, 8):
            for column in "ABC":
                cell = directive_styles[f"{column}{row}"]
                assert cell.fill.fill_type == "solid"
                assert cell.fill.fgColor.rgb == WHITE
                assert cell.border.left.style == "thin"
                assert cell.border.right.style == "thin"
                assert cell.border.top.style == "thin"
                assert cell.border.bottom.style == "thin"
        assert directive_styles["A8"].value.startswith("Footer moves")
    finally:
        workbook.close()
    return output_path


def main() -> None:
    """Generate both repeated-block sample workbooks and print their paths."""

    template = build_template()
    output = render_sample(template)
    print(f"Template: {template.relative_to(SAMPLES_DIR.parent)}")
    print(f"Output:   {output.relative_to(SAMPLES_DIR.parent)}")


if __name__ == "__main__":
    main()
