"""Render an eager Polars DataFrame through the optional adapter.

Run from the repository root with:

    uv run --extra polars python scratch/polars_example.py
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from excel_template_writer import normalize_context
from excel_template_writer.adapters.polars import polars_adapters
from excel_template_writer.xlsx import render_workbook

SCRATCH_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = SCRATCH_DIR / "polars_template.xlsx"
OUTPUT_PATH = SCRATCH_DIR / "polars_output.xlsx"


def _build_template() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Polars rows"
    sheet.append(["Description", "Quantity", "Amount", "Service date", "Discount"])
    sheet.append(
        [
            "{% for line in lines %}{{ line.description }}",
            "{{ line.quantity }}",
            "{{ line.amount }}",
            "{{ line.service_date }}",
            "{{ line.discount }}{% endfor %}",
        ]
    )

    border_side = Side(style="thin", color="FF8EA9C1")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF17365D")
        cell.border = border
    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor="FFEAF3F8")
        cell.border = border
    sheet["C2"].number_format = "$#,##0.00"
    sheet["D2"].number_format = "yyyy-mm-dd"
    sheet["E2"].number_format = "0%"
    sheet.freeze_panes = "A2"
    workbook.save(TEMPLATE_PATH)
    workbook.close()


def main() -> None:
    frame = pl.DataFrame(
        {
            "description": ["Consulting", "Support", "Training"],
            "quantity": [2, 4, 1],
            "amount": [1500.0, 350.0, 800.0],
            "service_date": [date(2026, 8, 1), date(2026, 8, 8), date(2026, 8, 15)],
            "discount": [0.10, float("nan"), None],
        }
    )
    adapters = polars_adapters()
    normalized = normalize_context({"lines": frame}, adapters=adapters).require()
    print(f"canonical rows: {normalized['lines']}")

    _build_template()
    render_workbook(
        TEMPLATE_PATH,
        OUTPUT_PATH,
        {"lines": frame},
        adapters=adapters,
    )

    rendered = load_workbook(OUTPUT_PATH)
    try:
        sheet = rendered["Polars rows"]
        assert [sheet.cell(row=row, column=1).value for row in range(2, 5)] == [
            "Consulting",
            "Support",
            "Training",
        ]
        assert sheet["C2"].value == 1500
        assert sheet["D3"].value == datetime(2026, 8, 8)
        assert sheet["E3"].value is None
        assert sheet["E4"].value is None
    finally:
        rendered.close()

    print(f"template: {TEMPLATE_PATH}")
    print(f"rendered: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
