"""Build a visual acceptance workbook for the current XLSX renderer.

Run from the repository root with:

    uv run python scratch/demo.py

The script deliberately creates the authored template before rendering it through the production
XLSX adapter. It then reopens the output and checks the workbook properties demonstrated on screen.
"""

from __future__ import annotations

import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from excel_template_writer.xlsx import render_workbook

SCRATCH_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = SCRATCH_DIR / "demo_template.xlsx"
OUTPUT_PATH = SCRATCH_DIR / "demo_output.xlsx"

# Eight-digit ARGB values make save/reload checks exact.
NAVY = "FF17365D"
BLUE = "FF5B9BD5"
LIGHT_BLUE = "FFDDEBF7"
PALE_BLUE = "FFEAF3F8"
GREEN = "FF70AD47"
LIGHT_GREEN = "FFE2F0D9"
RED = "FFC00000"
LIGHT_RED = "FFF4CCCC"
GOLD = "FFFFC000"
LIGHT_GOLD = "FFFFF2CC"
GRAY = "FF7F8C8D"
LIGHT_GRAY = "FFE7E6E6"
INK = "FF203040"
WHITE = "FFFFFFFF"
BORDER_COLOR = "FF8EA9C1"

THIN = Side(style="thin", color=BORDER_COLOR)
MEDIUM = Side(style="medium", color=NAVY)
DOUBLE = Side(style="double", color=NAVY)
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


DEMO_CONTEXT: dict[str, Any] = {
    "generated_for": "Formatting acceptance review",
    "report": {
        "issued_on": date(2026, 8, 13),
        "approved": True,
        "total": 12_345.67,
        "customer": "Acme Industries",
    },
    "supplied_text": "=THIS MUST REMAIN TEXT, NOT A FORMULA",
    "lines": [
        {
            "description": "Consulting",
            "quantity": 2,
            "unit_price": 1_500.0,
            "service_date": date(2026, 7, 2),
            "approved": True,
        },
        {
            "description": "Implementation",
            "quantity": 1,
            "unit_price": 4_250.0,
            "service_date": date(2026, 7, 15),
            "approved": False,
        },
        {
            "description": "Support",
            "quantity": 6,
            "unit_price": 350.0,
            "service_date": date(2026, 8, 1),
            "approved": True,
        },
    ],
    "cards": [
        {"title": "Launch", "owner": "Mina", "status": "On track", "budget": 75_000},
        {"title": "Renewal", "owner": "Jules", "status": "At risk", "budget": 42_500},
    ],
    "empty_items": [],
    "left_items": ["Left A", "Left B", "Left C"],
    "right_items": ["Right 1", "Right 2"],
    "account": {"active": True, "overdue": False},
    "groups": [
        {"name": "Hardware", "items": ["Laptop", "Monitor"], "total": 2},
        {"name": "Software", "items": ["Editor", "Database", "Support"], "total": 3},
    ],
}


def _atomic_save(workbook: WorkbookType, path: Path) -> None:
    """Save a generated fixture without leaving a partial workbook behind."""

    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(suffix=".xlsx", dir=path.parent)
    os.close(handle)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        load_workbook(temporary_path, read_only=True, data_only=False).close()
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _paint(
    sheet: Worksheet,
    cell_range: str,
    *,
    fill: str,
    font_color: str = INK,
    bold: bool = False,
    size: int = 11,
    horizontal: str = "left",
    border: Border = GRID_BORDER,
    wrap: bool = True,
) -> None:
    """Apply direct formatting to every cell in a range, including blank cells."""

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in sheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Aptos", size=size, bold=bold, color=font_color)
            cell.border = border
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical="center",
                wrap_text=wrap,
            )


def _merge_band(
    sheet: Worksheet,
    cell_range: str,
    value: str,
    *,
    fill: str,
    font_color: str = INK,
    bold: bool = False,
    size: int = 11,
    horizontal: str = "left",
) -> None:
    min_col, min_row, _, _ = range_boundaries(cell_range)
    sheet.merge_cells(cell_range)
    sheet.cell(min_row, min_col, value)
    _paint(
        sheet,
        cell_range,
        fill=fill,
        font_color=font_color,
        bold=bold,
        size=size,
        horizontal=horizontal,
    )


def _new_sheet(
    workbook: WorkbookType,
    title: str,
    description: str,
    *,
    widths: tuple[float, ...] = (34, 15, 18, 18, 17, 20),
) -> Worksheet:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    _merge_band(
        sheet,
        "A1:F1",
        title,
        fill=NAVY,
        font_color=WHITE,
        bold=True,
        size=18,
    )
    sheet.row_dimensions[1].height = 32
    _merge_band(sheet, "A2:F2", description, fill=PALE_BLUE, font_color=INK, size=10)
    sheet["A2"].font = Font(name="Aptos", size=10, italic=True, color=INK)
    sheet.row_dimensions[2].height = 38
    sheet.row_dimensions[3].height = 8
    return sheet


def _header(sheet: Worksheet, row: int, labels: tuple[str, ...]) -> None:
    for column, label in enumerate(labels, start=1):
        sheet.cell(row, column, label)
    _paint(
        sheet,
        f"A{row}:F{row}",
        fill=NAVY,
        font_color=WHITE,
        bold=True,
        horizontal="center",
    )
    sheet.row_dimensions[row].height = 24


def _build_start_here(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "START HERE",
        "Open this rendered workbook beside demo_template.xlsx. Each numbered sheet isolates one "
        "formatting or layout rule and states the expected result.",
        widths=(7, 24, 23, 23, 23, 23),
    )
    sheet.freeze_panes = "A9"
    _merge_band(
        sheet,
        "A4:F4",
        "How to inspect the demo",
        fill=BLUE,
        font_color=WHITE,
        bold=True,
        size=12,
    )
    instructions = (
        "1. Compare the same sheet in the template and output.",
        "2. Template tags must disappear; runtime values must keep native Excel types.",
        "3. Repeated rows and blocks must look like exact copies of the authored source block.",
    )
    for row, instruction in enumerate(instructions, start=5):
        sheet.cell(row, 1, row - 4)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        sheet.cell(row, 2, instruction)
        _paint(sheet, f"A{row}:F{row}", fill="FFFFFFFF")
        sheet.cell(row, 1).font = Font(name="Aptos", bold=True, color=NAVY)
        sheet.row_dimensions[row].height = 24

    _header(sheet, 9, ("#", "Sheet", "Primary check", "Secondary check", "Layout check", "Result"))
    checks = (
        (1, "Styled Table", "3 blue rows", "gold blank cells", "footer moves", "visually inspect"),
        (
            2,
            "Repeated Cards",
            "2 styled cards",
            "currency format",
            "merges repeat",
            "visually inspect",
        ),
        (3, "Empty Repeat", "1 green row", "blank keeps style", "footer stays", "visually inspect"),
        (
            4,
            "Cell Lanes",
            "left grows by 3",
            "right grows by 2",
            "middle stays put",
            "visually inspect",
        ),
        (5, "Conditions", "blue ACTIVE", "green CURRENT", "branches compact", "visually inspect"),
        (6, "Scalar Types", "date/number/bool", "cell formats", "no expansion", "visually inspect"),
        (
            7,
            "Nested Groups",
            "2 group headers",
            "5 item rows",
            "nested heights",
            "visually inspect",
        ),
    )
    for row, values in enumerate(checks, start=10):
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
        _paint(sheet, f"A{row}:F{row}", fill="FFFFFFFF" if row % 2 else PALE_BLUE)
        sheet.cell(row, 1).font = Font(name="Aptos", bold=True, color=NAVY)
        sheet.row_dimensions[row].height = 25

    _merge_band(
        sheet,
        "A18:F18",
        "Generated for: {{ generated_for }}",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    sheet.row_dimensions[18].height = 26


def _build_styled_table(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "1 Styled Table",
        "One authored body row becomes three complete rows. The source cell owns every style; "
        "the renderer does not infer formatting from the runtime value.",
    )
    _header(
        sheet, 5, ("Description", "Qty", "Unit price", "Service date", "Styled blank", "Approved")
    )

    sheet["A6"] = "{% for line in lines %}{{ line.description }}"
    sheet["B6"] = "{{ line.quantity }}"
    sheet["C6"] = "{{ line.unit_price }}"
    sheet["D6"] = "{{ line.service_date }}"
    sheet["E6"] = None
    sheet["F6"] = "{{ line.approved }}{% endfor %}"
    _paint(sheet, "A6:F6", fill=LIGHT_BLUE)
    sheet["A6"].font = Font(name="Aptos", bold=True, color=NAVY)
    sheet["B6"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["C6"].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    sheet["C6"].alignment = Alignment(horizontal="right", vertical="center")
    sheet["D6"].number_format = "yyyy-mm-dd"
    sheet["D6"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["E6"].fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    sheet["E6"].border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=DOUBLE)
    sheet["E6"].protection = Protection(locked=False)
    sheet["F6"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[6].height = 36
    sheet.row_dimensions[6].outlineLevel = 1

    _merge_band(
        sheet,
        "A7:F7",
        "GOLD FOOTER — must move below all three rendered rows",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    sheet.row_dimensions[7].height = 27


def _build_cards(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "2 Repeated Cards",
        "A two-row block repeats. Its merged navy title, detail row, row heights, borders, and "
        "currency number format must be copied as one unit.",
    )
    sheet.merge_cells("A5:F5")
    sheet["A5"] = "{% for card in cards %}{{ card.title }}"
    _paint(sheet, "A5:F5", fill=NAVY, font_color=WHITE, bold=True, size=14)
    sheet.row_dimensions[5].height = 30

    sheet["A6"] = "Owner"
    sheet["B6"] = "{{ card.owner }}"
    sheet["C6"] = "Status"
    sheet["D6"] = "{{ card.status }}"
    sheet["E6"] = "Budget"
    sheet["F6"] = "{{ card.budget }}{% endfor %}"
    _paint(sheet, "A6:F6", fill=PALE_BLUE)
    for coordinate in ("A6", "C6", "E6"):
        sheet[coordinate].font = Font(name="Aptos", bold=True, color=NAVY)
    sheet["F6"].number_format = '$#,##0;[Red]-$#,##0;"-"'
    sheet["F6"].alignment = Alignment(horizontal="right", vertical="center")
    sheet.row_dimensions[6].height = 24

    _merge_band(
        sheet,
        "A7:F7",
        "GOLD FOOTER — expected at row 9 after two cards",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    sheet.row_dimensions[7].height = 27


def _build_empty_repeat(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "3 Empty Repeat",
        "The collection is empty, so exactly one formatted placeholder row remains. Dynamic "
        "item values become blank; static text and blank-cell presentation remain.",
    )
    _header(sheet, 5, ("Description", "Styled blank", "State", "Qty", "Amount", "Ready"))
    sheet["A6"] = "{% for item in empty_items %}{{ item.description }}"
    sheet["B6"] = None
    sheet["C6"] = "FORMATTED PLACEHOLDER"
    sheet["D6"] = "{{ item.quantity }}"
    sheet["E6"] = "{{ item.amount }}"
    sheet["F6"] = "{{ item.ready }}{% endfor %}"
    _paint(sheet, "A6:F6", fill=LIGHT_GREEN)
    sheet["B6"].fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    sheet["B6"].border = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=DOUBLE)
    sheet["C6"].font = Font(name="Aptos", bold=True, color=GREEN)
    sheet["D6"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["E6"].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    sheet.row_dimensions[6].height = 32

    _merge_band(
        sheet,
        "A7:F7",
        "Footer remains on row 7 because the empty repeat keeps one source-sized instance",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    sheet.row_dimensions[7].height = 27


def _build_cell_lanes(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "4 Cell Lanes",
        'Two shift="cells" blocks grow independently. Only cells inside each two-column lane '
        "move; the gray middle columns stay at their authored coordinates.",
        widths=(24, 16, 24, 18, 24, 16),
    )
    _header(sheet, 4, ("Left lane", "Marker", "Stationary", "Stationary", "Right lane", "Marker"))
    sheet["A5"] = '{% for item in left_items shift="cells" %}{{ item }}'
    sheet["B5"] = "LEFT •{% endfor %}"
    sheet["C5"] = "STAYS AT C5"
    sheet["D5"] = "STAYS AT D5"
    sheet["E5"] = '{% for item in right_items shift="cells" %}{{ item }}'
    sheet["F5"] = "RIGHT ✓{% endfor %}"
    _paint(sheet, "A5:B5", fill=LIGHT_BLUE, bold=True)
    _paint(sheet, "C5:D5", fill=LIGHT_GRAY, bold=True, horizontal="center")
    _paint(sheet, "E5:F5", fill=LIGHT_GREEN, bold=True)
    # A cell-shift lane cannot own a worksheet-wide custom row height.
    sheet.row_dimensions[5].height = None

    sheet["A6"] = "AFTER LEFT — expected at A8"
    sheet["B6"] = "moves with left"
    sheet["C6"] = "STAYS AT C6"
    sheet["D6"] = "STAYS AT D6"
    sheet["E6"] = "AFTER RIGHT — expected at E7"
    sheet["F6"] = "moves with right"
    _paint(sheet, "A6:B6", fill=LIGHT_GOLD, bold=True)
    _paint(sheet, "C6:D6", fill=LIGHT_GRAY, bold=True, horizontal="center")
    _paint(sheet, "E6:F6", fill=LIGHT_GOLD, bold=True)
    sheet.row_dimensions[6].height = 28


def _build_conditions(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "5 Conditions",
        "Each if/else branch has deliberately different formatting. The selected rows compact "
        "upward and keep the formatting of the chosen source branch.",
    )
    sheet["A5"] = "{% if account.active %}Account"
    sheet["F5"] = "ACTIVE{% else %}"
    sheet["A6"] = "Account"
    sheet["F6"] = "INACTIVE{% endif %}"
    _paint(sheet, "A5:F5", fill=LIGHT_BLUE, font_color=NAVY, bold=True)
    _paint(sheet, "A6:F6", fill=LIGHT_GRAY, font_color=GRAY, bold=True)
    sheet.row_dimensions[5].height = 30
    sheet.row_dimensions[6].height = 26

    sheet["A7"] = "{% if account.overdue %}Payment"
    sheet["F7"] = "OVERDUE{% else %}"
    sheet["A8"] = "Payment"
    sheet["F8"] = "CURRENT{% endif %}"
    _paint(sheet, "A7:F7", fill=LIGHT_RED, font_color=RED, bold=True)
    _paint(sheet, "A8:F8", fill=LIGHT_GREEN, font_color=GREEN, bold=True)
    sheet.row_dimensions[7].height = 26
    sheet.row_dimensions[8].height = 30

    _merge_band(
        sheet,
        "A9:F9",
        "GOLD FOOTER — expected at row 7 after the two unselected branches disappear",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    sheet.row_dimensions[9].height = 27


def _build_scalars(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "6 Scalar Types",
        "Sole-expression cells preserve native Python/Excel types. Mixed-content cells become "
        "text, and source number formats remain authoritative.",
        widths=(25, 34, 22, 19, 19, 19),
    )
    _header(
        sheet,
        5,
        ("Case", "Template / rendered value", "Expected", "Type", "Format", "Visual check"),
    )
    rows = (
        ("Date", "{{ report.issued_on }}", "2026-08-13", "date", "yyyy-mm-dd", "date format"),
        ("Number", "{{ report.total }}", "12,345.67", "number", "currency", "right aligned"),
        ("Boolean", "{{ report.approved }}", "TRUE", "boolean", "General", "centered"),
        (
            "Mixed text",
            "Prepared for {{ report.customer }}",
            "Prepared for Acme Industries",
            "string",
            "General",
            "wraps",
        ),
        ("Equals text", "{{ supplied_text }}", "starts with =", "string", "Text", "not a formula"),
    )
    for row, values in enumerate(rows, start=6):
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
        _paint(sheet, f"A{row}:F{row}", fill="FFFFFFFF" if row % 2 else PALE_BLUE)
        sheet.cell(row, 1).font = Font(name="Aptos", bold=True, color=NAVY)
        sheet.row_dimensions[row].height = 30
    sheet["B6"].number_format = "yyyy-mm-dd"
    sheet["B6"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["B7"].number_format = '$#,##0.00;[Red]-$#,##0.00;"-"'
    sheet["B7"].alignment = Alignment(horizontal="right", vertical="center")
    sheet["B8"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["B10"].number_format = "@"


def _build_nested_groups(workbook: WorkbookType) -> None:
    sheet = _new_sheet(
        workbook,
        "7 Nested Groups",
        "The outer group repeat contains an inner item repeat. Measurement occurs before "
        "placement, so each group receives the height required by its own items.",
    )
    _header(sheet, 4, ("Group / item", "", "", "", "Kind", "Count"))
    sheet.merge_cells("A5:F5")
    sheet["A5"] = "{% for group in groups %}{{ group.name }}"
    _paint(sheet, "A5:F5", fill=NAVY, font_color=WHITE, bold=True, size=13)
    sheet.row_dimensions[5].height = 28

    sheet["A6"] = "{% for item in group.items %}{{ item }}"
    sheet["E6"] = "ITEM"
    sheet["F6"] = "•{% endfor %}"
    _paint(sheet, "A6:F6", fill=LIGHT_BLUE)
    sheet.row_dimensions[6].height = 23

    sheet["A7"] = "Group subtotal"
    sheet["E7"] = "COUNT"
    sheet["F7"] = "{{ group.total }}{% endfor %}"
    _paint(sheet, "A7:F7", fill=LIGHT_GRAY, bold=True)
    sheet["F7"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[7].height = 24

    _merge_band(
        sheet,
        "A8:F8",
        "GOLD FOOTER — expected at row 14 after both measured groups",
        fill=LIGHT_GOLD,
        bold=True,
        horizontal="center",
    )
    sheet.row_dimensions[8].height = 27


def build_demo_template(path: Path = TEMPLATE_PATH) -> Path:
    """Create the authored workbook containing visible tags and source formatting."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "Excel Template Writer formatting acceptance demo"
    workbook.properties.subject = "Visual source-to-output formatting verification"
    workbook.properties.creator = "excel-template-writer scratch demo"
    _build_start_here(workbook)
    _build_styled_table(workbook)
    _build_cards(workbook)
    _build_empty_repeat(workbook)
    _build_cell_lanes(workbook)
    _build_conditions(workbook)
    _build_scalars(workbook)
    _build_nested_groups(workbook)
    workbook.active = 0
    _atomic_save(workbook, path)
    return path


def _normalized_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    return value if isinstance(value, date) else None


def _assert_no_tags(workbook: WorkbookType) -> None:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert "{%" not in cell.value
                    assert "{{" not in cell.value


def _validate_output(template_path: Path, output_path: Path) -> None:
    """Reload both files and prove the workbook properties shown by the demo."""

    template = load_workbook(template_path, data_only=False)
    output = load_workbook(output_path, data_only=False)
    try:
        assert template.sheetnames == output.sheetnames
        assert any(
            "{%" in cell.value or "{{" in cell.value
            for sheet in template.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        )
        _assert_no_tags(output)

        start = output["START HERE"]
        assert start["A18"].value == "Generated for: Formatting acceptance review"

        table = output["1 Styled Table"]
        assert [table.cell(row, 1).value for row in range(6, 9)] == [
            "Consulting",
            "Implementation",
            "Support",
        ]
        assert [table.cell(row, 2).data_type for row in range(6, 9)] == ["n", "n", "n"]
        assert _normalized_date(table["D6"].value) == date(2026, 7, 2)
        assert table["F7"].value is False
        for row in range(6, 9):
            assert table.cell(row, 1).fill.fgColor.rgb == LIGHT_BLUE
            assert table.cell(row, 3).number_format == '$#,##0.00;[Red]-$#,##0.00;"-"'
            assert table.cell(row, 4).number_format == "yyyy-mm-dd"
            assert table.cell(row, 5).value is None
            assert table.cell(row, 5).fill.fgColor.rgb == LIGHT_GOLD
            assert table.cell(row, 5).border.bottom.style == "double"
            assert table.row_dimensions[row].height == 36
            assert table.row_dimensions[row].outlineLevel == 1
        assert "A9:F9" in table.merged_cells
        assert table["A9"].value.startswith("GOLD FOOTER")
        assert table.column_dimensions["A"].width == 34

        cards = output["2 Repeated Cards"]
        assert {"A5:F5", "A7:F7", "A9:F9"}.issubset(
            {str(item) for item in cards.merged_cells.ranges}
        )
        assert cards["A5"].value == "Launch"
        assert cards["A7"].value == "Renewal"
        assert cards["F8"].value == 42_500
        assert cards["F8"].number_format == '$#,##0;[Red]-$#,##0;"-"'
        assert [cards.row_dimensions[row].height for row in range(5, 9)] == [30, 24, 30, 24]

        empty = output["3 Empty Repeat"]
        assert empty["A6"].value is None
        assert empty["B6"].value is None
        assert empty["B6"].fill.fgColor.rgb == LIGHT_GOLD
        assert empty["C6"].value == "FORMATTED PLACEHOLDER"
        assert empty.row_dimensions[6].height == 32
        assert "A7:F7" in empty.merged_cells

        lanes = output["4 Cell Lanes"]
        assert [lanes.cell(row, 1).value for row in range(5, 8)] == [
            "Left A",
            "Left B",
            "Left C",
        ]
        assert lanes["A8"].value.startswith("AFTER LEFT")
        assert [lanes.cell(row, 5).value for row in range(5, 7)] == ["Right 1", "Right 2"]
        assert lanes["E7"].value.startswith("AFTER RIGHT")
        assert lanes["C5"].value == "STAYS AT C5"
        assert lanes["C6"].value == "STAYS AT C6"

        conditions = output["5 Conditions"]
        assert conditions["F5"].value == "ACTIVE"
        assert conditions["A5"].fill.fgColor.rgb == LIGHT_BLUE
        assert conditions["F6"].value == "CURRENT"
        assert conditions["A6"].fill.fgColor.rgb == LIGHT_GREEN
        assert conditions.row_dimensions[5].height == 30
        assert conditions.row_dimensions[6].height == 30
        assert "A7:F7" in conditions.merged_cells

        scalars = output["6 Scalar Types"]
        assert _normalized_date(scalars["B6"].value) == date(2026, 8, 13)
        assert scalars["B6"].data_type == "d"
        assert scalars["B7"].value == 12_345.67
        assert scalars["B7"].data_type == "n"
        assert scalars["B8"].value is True
        assert scalars["B8"].data_type == "b"
        assert scalars["B10"].value == "=THIS MUST REMAIN TEXT, NOT A FORMULA"
        assert scalars["B10"].data_type == "s"

        nested = output["7 Nested Groups"]
        assert nested["A5"].value == "Hardware"
        assert nested["A6"].value == "Laptop"
        assert nested["A7"].value == "Monitor"
        assert nested["A9"].value == "Software"
        assert nested["A12"].value == "Support"
        assert {"A5:F5", "A9:F9", "A14:F14"}.issubset(
            {str(item) for item in nested.merged_cells.ranges}
        )
        assert nested.row_dimensions[5].height == 28
        assert nested.row_dimensions[9].height == 28
    finally:
        template.close()
        output.close()


def render_demo(
    template_path: Path = TEMPLATE_PATH,
    output_path: Path = OUTPUT_PATH,
    context: dict[str, Any] | None = None,
) -> Path:
    """Render through the production adapter and verify the visual acceptance contract."""

    render_workbook(template_path, output_path, DEMO_CONTEXT if context is None else context)
    _validate_output(template_path, output_path)
    return output_path


def main() -> None:
    template_path = build_demo_template()
    output_path = render_demo(template_path)
    print(f"Template: {template_path}")
    print(f"Rendered: {output_path}")
    print("Validated after reload: values, types, styles, styled blanks, dimensions, and merges.")


if __name__ == "__main__":
    main()
