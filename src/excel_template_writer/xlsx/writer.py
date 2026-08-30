"""Apply complete render plans to a newly constructed XLSX workbook."""

from __future__ import annotations

import os
import posixpath
import tempfile
from copy import copy, deepcopy
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, OneCellAnchor, TwoCellAnchor
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.worksheet.worksheet import Worksheet

from excel_template_writer.diagnostics import TemplateRenderError
from excel_template_writer.limits import ResourceLimits
from excel_template_writer.model import Coordinate
from excel_template_writer.render import RenderPlan
from excel_template_writer.xlsx.model import (
    CellPresentation,
    ChartSnapshot,
    ColumnPresentation,
    DimensionPresentation,
    ImageSnapshot,
    RowPresentation,
    SheetFeaturePlan,
    SheetSnapshot,
    TextShapeSnapshot,
    WorkbookSnapshot,
)
from excel_template_writer.xlsx.package_limits import inspect_xlsx_package

_CHART_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWING_MAIN_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/main"
_OFFICE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_DRAWING_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_PACKAGE_RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_TYPES_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/content-types"
_DRAWING_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
)
_DRAWING_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.drawing+xml"

ElementTree.register_namespace("xdr", _DRAWING_NAMESPACE)
ElementTree.register_namespace("a", _DRAWING_MAIN_NAMESPACE)
ElementTree.register_namespace("r", _OFFICE_RELATIONSHIP_NAMESPACE)


class _PreservedImage(Image):
    def __init__(self, data: bytes, image_format: str, anchor: Any) -> None:
        """Create an openpyxl-compatible image without decoding its media bytes.

        Args:
            data: Exact embedded media payload from the source XLSX package.
            image_format: Supported file extension used for the output media part.
            anchor: Detached DrawingML anchor containing the source picture frame.
        """

        self._payload = data
        self.format = image_format
        self.anchor = anchor

    def _data(self) -> bytes:
        """Return the exact source media payload for package serialization."""

        return self._payload


def _apply_dimension_style(
    destination: RowDimension | ColumnDimension,
    source: DimensionPresentation,
) -> None:
    """Copy formatting shared by row and column dimensions.

    Args:
        destination: Mutable openpyxl dimension receiving formatting.
        source: Detached source dimension presentation.
    """

    destination.font = copy(source.font)
    destination.fill = copy(source.fill)
    destination.border = copy(source.border)
    destination.alignment = copy(source.alignment)
    destination.number_format = source.number_format
    destination.protection = copy(source.protection)


def _apply_row(destination: RowDimension, source: RowPresentation) -> None:
    """Apply detached row presentation to an output row dimension.

    Args:
        destination: Mutable output row dimension.
        source: Detached source row presentation.
    """

    destination.height = source.height
    destination.hidden = source.hidden
    destination.outlineLevel = source.outline_level
    destination.collapsed = source.collapsed
    destination.thickTop = source.thick_top
    destination.thickBot = source.thick_bottom
    _apply_dimension_style(destination, source)


def _apply_column(destination: ColumnDimension, source: ColumnPresentation) -> None:
    """Apply detached column presentation to an output column dimension.

    Args:
        destination: Mutable output column dimension.
        source: Detached source column presentation.
    """

    destination.width = source.width
    destination.hidden = source.hidden
    destination.outlineLevel = source.outline_level
    destination.collapsed = source.collapsed
    destination.bestFit = source.best_fit
    _apply_dimension_style(destination, source)


def _apply_cell(destination: Cell, source: CellPresentation, value: object) -> None:
    """Write one planned value and copy its direct source presentation.

    Args:
        destination: Mutable output cell.
        source: Detached source cell presentation.
        value: Evaluated planned value to write.
    """

    destination.value = value
    if isinstance(value, str) and value.startswith("=") and not source.is_formula:
        destination.data_type = "s"
    destination.font = copy(source.font)
    destination.fill = copy(source.fill)
    destination.border = copy(source.border)
    destination.alignment = copy(source.alignment)
    destination.number_format = source.number_format
    destination.protection = copy(source.protection)
    destination.quotePrefix = source.quote_prefix
    destination.hyperlink = copy(source.hyperlink)
    destination.comment = copy(source.comment)


def _is_identity_plan(source: SheetSnapshot, plan: RenderPlan) -> bool:
    """Return whether a plan leaves every supported coordinate unchanged.

    Args:
        source: Source worksheet snapshot.
        plan: Completed render plan.

    Returns:
        ``True`` when cell and merge coordinates are identical.
    """

    cell_mappings = [(cell.source_coordinate, cell.coordinate) for cell in plan.cells]
    source_merges = set(source.template.merged_ranges)
    planned_merges = {
        merge.rectangle for merge in plan.merges if merge.source_rectangle == merge.rectangle
    }
    return (
        len(cell_mappings) == len(source.cells)
        and all(
            source_coordinate == destination for source_coordinate, destination in cell_mappings
        )
        and source_merges == planned_merges
    )


def _apply_drawing_anchor(anchor: Any, planned_coordinates: tuple[Coordinate, ...]) -> None:
    """Apply one prevalidated translation to a detached drawing anchor.

    Args:
        anchor: Mutable openpyxl drawing anchor.
        planned_coordinates: Destination coordinates for its cell anchor markers.

    Raises:
        RuntimeError: If the validated anchor and feature plan disagree.
    """

    if isinstance(anchor, AbsoluteAnchor):
        if planned_coordinates:
            raise RuntimeError("absolute drawing anchor unexpectedly has cell markers")
        return
    if isinstance(anchor, OneCellAnchor):
        markers = (anchor._from,)
    elif isinstance(anchor, TwoCellAnchor):
        markers = (anchor._from, anchor.to)
    else:
        raise RuntimeError("unsupported drawing anchor reached the workbook writer")
    if len(markers) != len(planned_coordinates):
        raise RuntimeError("drawing anchor marker count does not match its feature plan")
    for marker, coordinate in zip(markers, planned_coordinates, strict=True):
        marker.row = coordinate.row - 1
        marker.col = coordinate.column - 1


def _drawing_relationship_id(anchor: ElementTree.Element) -> str | None:
    """Return the chart or image relationship ID used by one drawing anchor.

    Args:
        anchor: Serialized anchor element from an output drawing part.

    Returns:
        Embedded relationship ID, or ``None`` for an unexpected drawing object.
    """

    chart = anchor.find(f".//{{{_CHART_NAMESPACE}}}chart")
    if chart is not None:
        return chart.get(f"{{{_OFFICE_RELATIONSHIP_NAMESPACE}}}id")
    blip = anchor.find(f".//{{{_DRAWING_MAIN_NAMESPACE}}}blip")
    if blip is not None:
        return blip.get(f"{{{_OFFICE_RELATIONSHIP_NAMESPACE}}}embed")
    return None


def _qualified_name(namespace: str, local_name: str) -> str:
    """Return an ElementTree expanded XML name.

    Args:
        namespace: XML namespace URI.
        local_name: Local element or attribute name.

    Returns:
        Expanded ``{namespace}local_name`` value.
    """

    return f"{{{namespace}}}{local_name}"


def _generated_relationship_ids(source: SheetSnapshot) -> dict[int, str]:
    """Map source drawing indexes to IDs assigned by openpyxl on this sheet.

    Args:
        source: Worksheet snapshot containing drawings in authored stacking order.

    Returns:
        Source indexes for generated charts and images mapped to output relationship IDs.
    """

    chart_count = sum(isinstance(drawing, ChartSnapshot) for drawing in source.drawings)
    chart_index = 0
    image_index = 0
    relationship_ids: dict[int, str] = {}
    for index, drawing in enumerate(source.drawings):
        if isinstance(drawing, ChartSnapshot):
            chart_index += 1
            relationship_ids[index] = f"rId{chart_index}"
        elif isinstance(drawing, ImageSnapshot):
            image_index += 1
            relationship_ids[index] = f"rId{chart_count + image_index}"
    return relationship_ids


def _apply_text_shape_anchor(
    source: TextShapeSnapshot,
    planned_coordinates: tuple[Coordinate, ...],
) -> ElementTree.Element:
    """Deserialize one preserved shape and apply its prevalidated translation.

    Args:
        source: Shape snapshot containing its source anchor XML.
        planned_coordinates: Destination cell markers from the completed feature plan.

    Returns:
        Mutable translated DrawingML anchor ready for package serialization.

    Raises:
        RuntimeError: If the snapshot and validated feature plan disagree.
    """

    anchor = ElementTree.fromstring(source.anchor_xml)
    if anchor.tag == _qualified_name(_DRAWING_NAMESPACE, "absoluteAnchor"):
        marker_names: tuple[str, ...] = ()
    elif anchor.tag == _qualified_name(_DRAWING_NAMESPACE, "oneCellAnchor"):
        marker_names = ("from",)
    elif anchor.tag == _qualified_name(_DRAWING_NAMESPACE, "twoCellAnchor"):
        marker_names = ("from", "to")
    else:
        raise RuntimeError("unsupported text-shape anchor reached the workbook writer")
    if len(marker_names) != len(planned_coordinates):
        raise RuntimeError("text-shape marker count does not match its feature plan")
    for marker_name, coordinate in zip(marker_names, planned_coordinates, strict=True):
        marker = anchor.find(_qualified_name(_DRAWING_NAMESPACE, marker_name))
        if marker is None:
            raise RuntimeError("text-shape anchor marker is missing")
        row = marker.find(_qualified_name(_DRAWING_NAMESPACE, "row"))
        column = marker.find(_qualified_name(_DRAWING_NAMESPACE, "col"))
        if row is None or column is None:
            raise RuntimeError("text-shape anchor marker coordinates are missing")
        row.text = str(coordinate.row - 1)
        column.text = str(coordinate.column - 1)
    return anchor


def _renumber_drawing_objects(root: ElementTree.Element) -> None:
    """Assign package-local unique non-visual IDs in final stacking order.

    Args:
        root: Worksheet drawing root whose object IDs are normalized in place.
    """

    next_id = 1
    non_visual_tag = _qualified_name(_DRAWING_NAMESPACE, "cNvPr")
    for anchor in root:
        for properties in anchor.iter(non_visual_tag):
            properties.set("id", str(next_id))
            next_id += 1


def _worksheet_relationship_part(sheet_part: str) -> str:
    """Return the relationship-part path owned by one worksheet part.

    Args:
        sheet_part: Package-relative worksheet XML path.

    Returns:
        Package-relative path to the worksheet's relationship part.
    """

    directory, filename = posixpath.split(sheet_part)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _normalized_part_target(source_part: str, target: str) -> str:
    """Resolve one package relationship target to a package-relative path.

    Args:
        source_part: Package-relative path of the relationship owner.
        target: Absolute or relative OPC relationship target.

    Returns:
        Normalized package-relative target without a leading slash.
    """

    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _next_relationship_id(root: ElementTree.Element) -> str:
    """Allocate an unused conventional relationship ID.

    Args:
        root: OPC relationships root containing zero or more IDs.

    Returns:
        First unused conventional ``rIdN`` value.
    """

    used = {element.get("Id") for element in root}
    index = 1
    while f"rId{index}" in used:
        index += 1
    return f"rId{index}"


def _append_worksheet_drawing(
    worksheet_xml: bytes,
    relationship_id: str,
) -> bytes:
    """Attach a newly created drawing relationship to a worksheet XML part.

    Args:
        worksheet_xml: Serialized worksheet part without a drawing element.
        relationship_id: Worksheet relationship ID for the new drawing part.

    Returns:
        Worksheet XML containing the schema-ordered drawing element.

    Raises:
        RuntimeError: If the worksheet already contains a drawing element.
    """

    root = ElementTree.fromstring(worksheet_xml)
    drawing_tag = _qualified_name(_SPREADSHEET_NAMESPACE, "drawing")
    if root.find(drawing_tag) is not None:
        raise RuntimeError("worksheet unexpectedly already contains a drawing element")
    drawing = ElementTree.Element(
        drawing_tag,
        {_qualified_name(_OFFICE_RELATIONSHIP_NAMESPACE, "id"): relationship_id},
    )
    following_tags = {
        _qualified_name(_SPREADSHEET_NAMESPACE, name)
        for name in (
            "legacyDrawing",
            "legacyDrawingHF",
            "picture",
            "oleObjects",
            "controls",
            "webPublishItems",
            "tableParts",
            "extLst",
        )
    }
    insertion_index = next(
        (index for index, child in enumerate(root) if child.tag in following_tags),
        len(root),
    )
    root.insert(insertion_index, drawing)
    return ElementTree.tostring(root, encoding="utf-8")


def _restore_drawing_order(
    path: Path,
    snapshot: WorkbookSnapshot,
    feature_plans: tuple[SheetFeaturePlan, ...],
) -> None:
    """Inject preserved shapes and restore source drawing stacking order.

    Args:
        path: Serialized temporary XLSX package to rewrite atomically.
        snapshot: Source workbook containing ordered drawing snapshots.
        feature_plans: Validated destination anchors for every worksheet drawing.

    Raises:
        RuntimeError: If serialized relationships or anchors disagree with the plans.
    """

    if not any(sheet.drawings for sheet in snapshot.sheets):
        return
    handle, repacked_name = tempfile.mkstemp(suffix=".xlsx", dir=path.parent)
    os.close(handle)
    repacked_path = Path(repacked_name)
    try:
        with ZipFile(path, "r") as archive:
            archive_names = frozenset(archive.namelist())
            modified_parts: dict[str, bytes] = {}
            new_parts: dict[str, bytes] = {}
            drawing_indexes = [
                int(posixpath.basename(name).removeprefix("drawing").removesuffix(".xml"))
                for name in archive_names
                if name.startswith("xl/drawings/drawing")
                and name.endswith(".xml")
                and posixpath.basename(name).removeprefix("drawing").removesuffix(".xml").isdigit()
            ]
            next_drawing_index = max(drawing_indexes, default=0) + 1
            created_drawing_parts: list[str] = []

            for sheet_index, (sheet, feature_plan) in enumerate(
                zip(snapshot.sheets, feature_plans, strict=True),
                start=1,
            ):
                if not sheet.drawings:
                    continue
                sheet_part = f"xl/worksheets/sheet{sheet_index}.xml"
                relationships_part = _worksheet_relationship_part(sheet_part)
                if relationships_part in archive_names:
                    relationships_root = ElementTree.fromstring(archive.read(relationships_part))
                else:
                    relationships_root = ElementTree.Element(
                        _qualified_name(_PACKAGE_RELATIONSHIP_NAMESPACE, "Relationships")
                    )
                drawing_relationships = [
                    relationship
                    for relationship in relationships_root
                    if relationship.get("Type") == _DRAWING_RELATIONSHIP
                ]
                if len(drawing_relationships) > 1:
                    raise RuntimeError("worksheet has multiple drawing relationships")

                if drawing_relationships:
                    relationship = drawing_relationships[0]
                    target = relationship.get("Target")
                    if target is None:
                        raise RuntimeError("worksheet drawing relationship has no target")
                    drawing_part = _normalized_part_target(sheet_part, target)
                    if drawing_part not in archive_names:
                        raise RuntimeError("serialized worksheet drawing part is missing")
                    drawing_root = ElementTree.fromstring(archive.read(drawing_part))
                else:
                    if any(
                        isinstance(drawing, (ChartSnapshot, ImageSnapshot))
                        for drawing in sheet.drawings
                    ):
                        raise RuntimeError("serialized worksheet drawing relationship is missing")
                    drawing_part = f"xl/drawings/drawing{next_drawing_index}.xml"
                    next_drawing_index += 1
                    created_drawing_parts.append(drawing_part)
                    drawing_root = ElementTree.Element(_qualified_name(_DRAWING_NAMESPACE, "wsDr"))
                    relationship_id = _next_relationship_id(relationships_root)
                    ElementTree.SubElement(
                        relationships_root,
                        _qualified_name(
                            _PACKAGE_RELATIONSHIP_NAMESPACE,
                            "Relationship",
                        ),
                        {
                            "Id": relationship_id,
                            "Type": _DRAWING_RELATIONSHIP,
                            "Target": f"/{drawing_part}",
                        },
                    )
                    serialized_relationships = ElementTree.tostring(
                        relationships_root,
                        encoding="utf-8",
                    )
                    if relationships_part in archive_names:
                        modified_parts[relationships_part] = serialized_relationships
                    else:
                        new_parts[relationships_part] = serialized_relationships
                    modified_parts[sheet_part] = _append_worksheet_drawing(
                        archive.read(sheet_part),
                        relationship_id,
                    )

                generated_relationship_ids = _generated_relationship_ids(sheet)
                anchors_by_relationship: dict[str, ElementTree.Element] = {}
                for anchor in drawing_root:
                    relationship_id = _drawing_relationship_id(anchor)
                    if relationship_id is None or relationship_id in anchors_by_relationship:
                        raise RuntimeError("serialized worksheet drawing order is ambiguous")
                    anchors_by_relationship[relationship_id] = anchor
                if set(anchors_by_relationship) != set(generated_relationship_ids.values()):
                    raise RuntimeError(
                        "serialized worksheet drawings do not match their feature plan"
                    )

                ordered_anchors: list[ElementTree.Element] = []
                for index, (drawing, planned_drawing) in enumerate(
                    zip(sheet.drawings, feature_plan.drawings, strict=True)
                ):
                    if isinstance(drawing, TextShapeSnapshot):
                        ordered_anchors.append(
                            _apply_text_shape_anchor(
                                drawing,
                                planned_drawing.anchor_coordinates,
                            )
                        )
                    else:
                        ordered_anchors.append(
                            anchors_by_relationship[generated_relationship_ids[index]]
                        )
                drawing_root[:] = ordered_anchors
                _renumber_drawing_objects(drawing_root)
                serialized_drawing = ElementTree.tostring(
                    drawing_root,
                    encoding="utf-8",
                )
                if drawing_part in archive_names:
                    modified_parts[drawing_part] = serialized_drawing
                else:
                    new_parts[drawing_part] = serialized_drawing

            if created_drawing_parts:
                content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
                existing_overrides = {
                    element.get("PartName")
                    for element in content_types
                    if element.tag == _qualified_name(_CONTENT_TYPES_NAMESPACE, "Override")
                }
                for drawing_part in created_drawing_parts:
                    part_name = f"/{drawing_part}"
                    if part_name not in existing_overrides:
                        ElementTree.SubElement(
                            content_types,
                            _qualified_name(_CONTENT_TYPES_NAMESPACE, "Override"),
                            {
                                "PartName": part_name,
                                "ContentType": _DRAWING_CONTENT_TYPE,
                            },
                        )
                modified_parts["[Content_Types].xml"] = ElementTree.tostring(
                    content_types,
                    encoding="utf-8",
                )

            with ZipFile(repacked_path, "w") as destination:
                destination.comment = archive.comment
                for member in archive.infolist():
                    destination.writestr(
                        member,
                        modified_parts.get(member.filename, archive.read(member.filename)),
                    )
                for part_name, data in new_parts.items():
                    destination.writestr(part_name, data)
        os.replace(repacked_path, path)
    finally:
        repacked_path.unlink(missing_ok=True)


def _write_sheet(
    destination: Worksheet,
    source: SheetSnapshot,
    plan: RenderPlan,
    feature_plan: SheetFeaturePlan,
) -> None:
    """Apply one complete validated plan to a new worksheet.

    Args:
        destination: Empty mutable destination worksheet.
        source: Detached source worksheet state.
        plan: Complete adapter-neutral render plan.
        feature_plan: Validated adapter plan for charts and other workbook features.
    """

    destination.sheet_view.showGridLines = source.show_grid_lines
    destination.freeze_panes = source.freeze_panes
    destination.sheet_properties.tabColor = copy(source.tab_color)
    destination.sheet_format.defaultRowHeight = source.default_row_height
    destination.sheet_format.defaultColWidth = source.default_column_width

    for column, presentation in source.columns.items():
        _apply_column(destination.column_dimensions[get_column_letter(column)], presentation)
    for planned_row in plan.rows:
        if planned_row.source_row is None:
            continue
        presentation = source.rows.get(planned_row.source_row)
        if presentation is not None:
            _apply_row(destination.row_dimensions[planned_row.destination_row], presentation)

    for planned_cell in plan.cells:
        if planned_cell.source_coordinate in source.synthetic_drawing_anchor_cells:
            continue
        presentation = source.cells[planned_cell.source_coordinate]
        cell = destination.cell(planned_cell.coordinate.row, planned_cell.coordinate.column)
        _apply_cell(cell, presentation, planned_cell.value)

    for merge in plan.merges:
        destination.merge_cells(
            start_row=merge.rectangle.top,
            start_column=merge.rectangle.left,
            end_row=merge.rectangle.bottom,
            end_column=merge.rectangle.right,
        )

    for source_drawing, planned_drawing in zip(
        source.drawings,
        feature_plan.drawings,
        strict=True,
    ):
        if isinstance(source_drawing, ChartSnapshot):
            chart = deepcopy(source_drawing.chart)
            _apply_drawing_anchor(chart.anchor, planned_drawing.anchor_coordinates)
            destination.add_chart(chart)
        elif isinstance(source_drawing, ImageSnapshot):
            anchor = deepcopy(source_drawing.anchor)
            _apply_drawing_anchor(anchor, planned_drawing.anchor_coordinates)
            destination.add_image(
                _PreservedImage(
                    source_drawing.data,
                    source_drawing.image_format,
                    anchor,
                )
            )
        else:
            assert isinstance(source_drawing, TextShapeSnapshot)

    if _is_identity_plan(source, plan):
        destination.conditional_formatting = copy(source.conditional_formatting)
        destination.data_validations = copy(source.data_validations)
        destination.auto_filter = copy(source.auto_filter)


def write_workbook(
    snapshot: WorkbookSnapshot,
    plans: tuple[RenderPlan, ...],
    feature_plans: tuple[SheetFeaturePlan, ...],
    output_path: str | Path,
    *,
    limits: ResourceLimits,
) -> Path:
    """Write atomically and reopen the package before publishing it.

    Args:
        snapshot: Detached source workbook state.
        plans: One complete render plan per source worksheet.
        feature_plans: One validated workbook-feature plan per source worksheet.
        output_path: Destination path, which must differ from the template path.
        limits: Package ceilings checked before publication.

    Returns:
        Resolved destination path after successful verification.

    Raises:
        TemplateRenderError: If the serialized package exceeds resource limits.
    """

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties = copy(snapshot.properties)
    workbook.loaded_theme = snapshot.loaded_theme
    for sheet, plan, feature_plan in zip(
        snapshot.sheets,
        plans,
        feature_plans,
        strict=True,
    ):
        destination = workbook.create_sheet(sheet.template.name)
        _write_sheet(destination, sheet, plan, feature_plan)

    handle, temporary_name = tempfile.mkstemp(suffix=".xlsx", dir=path.parent)
    os.close(handle)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        _restore_drawing_order(temporary_path, snapshot, feature_plans)
        package_diagnostic = inspect_xlsx_package(
            temporary_path,
            limits,
            description="rendered XLSX package",
        )
        if package_diagnostic is not None:
            raise TemplateRenderError((package_diagnostic,))
        verified = load_workbook(temporary_path, read_only=True, data_only=False)
        verified.close()
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)
    return path
