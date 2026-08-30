"""Read an XLSX workbook into adapter-owned immutable snapshots."""

from __future__ import annotations

import posixpath
from collections.abc import Iterable
from copy import copy, deepcopy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, ScatterChart
from openpyxl.chart.chartspace import ChartSpace
from openpyxl.chart.reader import read_chart
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, OneCellAnchor, TwoCellAnchor
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from excel_template_writer.model import Coordinate, Rectangle, WorksheetTemplate
from excel_template_writer.xlsx.model import (
    CellPresentation,
    ChartSnapshot,
    ColumnPresentation,
    DimensionPresentation,
    DrawingSnapshot,
    ImageSnapshot,
    RowPresentation,
    SheetSnapshot,
    WorkbookSnapshot,
)

_CHART_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWING_MAIN_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/main"
_DRAWING_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DRAWING_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
)
_CHART_RELATIONSHIP = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"
_IMAGE_RELATIONSHIP = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
_OFFICE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_PACKAGE_RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_SUPPORTED_CHART_TYPES = (AreaChart, BarChart, LineChart, PieChart, ScatterChart)
_SUPPORTED_ANCHOR_TYPES = (AbsoluteAnchor, OneCellAnchor, TwoCellAnchor)
_SUPPORTED_IMAGE_FORMATS = frozenset({"jpeg", "jpg", "png"})


@dataclass(frozen=True)
class _PackageRelationship:
    relationship_type: str
    target: str
    is_external: bool


@dataclass(frozen=True)
class _DrawingPartProfile:
    drawings: tuple[DrawingSnapshot, ...]
    has_unsupported_objects: bool


def _has_supported_image_media(image_format: str, data: bytes) -> bool:
    """Return whether media has a supported extension and basic file signature.

    Args:
        image_format: Lowercase media-part extension.
        data: Embedded media payload.

    Returns:
        ``True`` for structurally recognizable PNG or JPEG media.
    """

    if image_format == "png":
        return data.startswith(b"\x89PNG\r\n\x1a\n")
    if image_format in {"jpeg", "jpg"}:
        return data.startswith(b"\xff\xd8") and data.endswith(b"\xff\xd9")
    return False


def _qualified_name(namespace: str, local_name: str) -> str:
    """Return one ElementTree expanded XML name.

    Args:
        namespace: XML namespace URI.
        local_name: Local element name.

    Returns:
        ElementTree expanded name in ``{namespace}local`` form.
    """

    return f"{{{namespace}}}{local_name}"


def _normalized_part_target(source_part: str, target: str) -> str:
    """Resolve one OOXML relationship target to a package-relative path.

    Args:
        source_part: Package-relative path of the relationship owner.
        target: Absolute or relative relationship target.

    Returns:
        Normalized package-relative target without a leading slash.
    """

    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _drawing_relationships(
    archive: ZipFile,
    drawing_name: str,
) -> dict[str, _PackageRelationship]:
    """Resolve relationships owned by one drawing part.

    Args:
        archive: Open XLSX ZIP archive.
        drawing_name: Package-relative drawing-part path.

    Returns:
        Relationship IDs mapped to their type, target, and external flag.
    """

    directory, filename = posixpath.split(drawing_name)
    relationships_name = posixpath.join(directory, "_rels", f"{filename}.rels")
    if relationships_name not in archive.namelist():
        return {}
    root = ElementTree.fromstring(archive.read(relationships_name))
    relationships: dict[str, _PackageRelationship] = {}
    for element in root:
        if element.tag != _qualified_name(_PACKAGE_RELATIONSHIP_NAMESPACE, "Relationship"):
            continue
        relationship_id = element.get("Id")
        relationship_type = element.get("Type")
        target = element.get("Target")
        is_external = element.get("TargetMode") == "External"
        if relationship_id is not None and relationship_type is not None and target is not None:
            relationships[relationship_id] = _PackageRelationship(
                relationship_type,
                target if is_external else _normalized_part_target(drawing_name, target),
                is_external,
            )
    return relationships


def _anchor_coordinates(anchor: Any) -> tuple[Coordinate, ...]:
    """Return the cell markers used by one supported drawing anchor.

    Args:
        anchor: Openpyxl absolute, one-cell, two-cell, or unsupported anchor.

    Returns:
        One-based worksheet coordinates for every cell-based marker.
    """

    if isinstance(anchor, AbsoluteAnchor):
        return ()
    if isinstance(anchor, OneCellAnchor):
        markers = (anchor._from,)
    elif isinstance(anchor, TwoCellAnchor):
        markers = (anchor._from, anchor.to)
    else:
        return ()
    return tuple(Coordinate(marker.row + 1, marker.col + 1) for marker in markers)


def _chart_references(chart: Any) -> tuple[str, ...]:
    """Serialize one detached chart and collect every chart-formula reference.

    Args:
        chart: Detached openpyxl chart object.

    Returns:
        Formula strings in serialized chart order.
    """

    tree = deepcopy(chart)._write()
    return tuple(
        element.text
        for element in tree.iter()
        if element.tag.rsplit("}", 1)[-1] == "f" and element.text
    )


def _parse_anchor(element: ElementTree.Element) -> Any | None:
    """Parse one supported DrawingML anchor element.

    Args:
        element: Raw anchor element from a worksheet drawing part.

    Returns:
        Parsed openpyxl anchor, or ``None`` when the type or content is unsupported.
    """

    anchor_type = {
        _qualified_name(_DRAWING_NAMESPACE, "absoluteAnchor"): AbsoluteAnchor,
        _qualified_name(_DRAWING_NAMESPACE, "oneCellAnchor"): OneCellAnchor,
        _qualified_name(_DRAWING_NAMESPACE, "twoCellAnchor"): TwoCellAnchor,
    }.get(element.tag)
    if anchor_type is None:
        return None
    try:
        return anchor_type.from_tree(element)
    except (TypeError, ValueError):
        return None


def _read_chart_part(
    archive: ZipFile,
    chart_part: str,
    anchor: Any,
) -> ChartSnapshot:
    """Read one chart part and bind its exact source drawing anchor.

    Args:
        archive: XLSX package containing the chart.
        chart_part: Package-relative chart XML path.
        anchor: Parsed source drawing anchor.

    Returns:
        Detached validated-chart input snapshot.
    """

    chart_space = ChartSpace.from_tree(ElementTree.fromstring(archive.read(chart_part)))
    chart = read_chart(chart_space)
    chart.style = chart_space.style
    chart.roundedCorners = chart_space.roundedCorners
    detached_anchor = deepcopy(anchor)
    detached_anchor.graphicFrame = None
    chart.anchor = detached_anchor
    return ChartSnapshot(
        chart=chart,
        chart_type=type(chart).__name__,
        anchor_coordinates=_anchor_coordinates(detached_anchor),
        references=_chart_references(chart),
        has_supported_type=type(chart) in _SUPPORTED_CHART_TYPES,
        has_supported_anchor=type(detached_anchor) in _SUPPORTED_ANCHOR_TYPES,
        is_combined=len(chart._charts) != 1,
        is_pivot=chart.pivotSource is not None,
    )


def _drawing_profiles(path: Path) -> dict[str, _DrawingPartProfile]:
    """Inspect and snapshot ordered supported worksheet drawings.

    Args:
        path: XLSX package to inspect.

    Returns:
        Drawing-part paths mapped to ordered drawings and unsupported-object state.
    """

    profiles: dict[str, _DrawingPartProfile] = {}
    with ZipFile(path) as archive:
        archive_names = frozenset(archive.namelist())
        drawing_names = (
            name
            for name in archive_names
            if name.startswith("xl/drawings/drawing") and name.endswith(".xml")
        )
        for name in drawing_names:
            root = ElementTree.fromstring(archive.read(name))
            relationships = _drawing_relationships(archive, name)
            drawings: list[DrawingSnapshot] = []
            used_relationship_ids: set[str] = set()
            has_unsupported = root.tag != _qualified_name(_DRAWING_NAMESPACE, "wsDr")
            anchors = list(root)
            if not anchors:
                has_unsupported = True
            for anchor_element in anchors:
                anchor = _parse_anchor(anchor_element)
                if anchor is None:
                    has_unsupported = True
                    continue
                object_children: list[ElementTree.Element] = []
                for child in anchor_element:
                    if child.tag in {
                        _qualified_name(_DRAWING_NAMESPACE, "clientData"),
                        _qualified_name(_DRAWING_NAMESPACE, "ext"),
                        _qualified_name(_DRAWING_NAMESPACE, "from"),
                        _qualified_name(_DRAWING_NAMESPACE, "pos"),
                        _qualified_name(_DRAWING_NAMESPACE, "to"),
                    }:
                        continue
                    object_children.append(child)
                if len(object_children) != 1:
                    has_unsupported = True
                    continue
                drawing_object = object_children[0]
                if drawing_object.tag == _qualified_name(_DRAWING_NAMESPACE, "graphicFrame"):
                    chart_elements = [
                        descendant
                        for descendant in drawing_object.iter()
                        if descendant.tag == _qualified_name(_CHART_NAMESPACE, "chart")
                    ]
                    if len(chart_elements) != 1:
                        has_unsupported = True
                        continue
                    relationship_id = chart_elements[0].get(
                        _qualified_name(_OFFICE_RELATIONSHIP_NAMESPACE, "id")
                    )
                    relationship = relationships.get(relationship_id or "")
                    if (
                        relationship_id is None
                        or relationship is None
                        or relationship.relationship_type != _CHART_RELATIONSHIP
                        or relationship.is_external
                        or relationship.target not in archive_names
                    ):
                        has_unsupported = True
                        continue
                    used_relationship_ids.add(relationship_id)
                    try:
                        drawings.append(_read_chart_part(archive, relationship.target, anchor))
                    except (TypeError, ValueError, KeyError, IndexError):
                        has_unsupported = True
                    continue
                if drawing_object.tag == _qualified_name(_DRAWING_NAMESPACE, "pic"):
                    blips = [
                        descendant
                        for descendant in drawing_object.iter()
                        if descendant.tag == _qualified_name(_DRAWING_MAIN_NAMESPACE, "blip")
                    ]
                    if len(blips) != 1:
                        has_unsupported = True
                        continue
                    embed_id = blips[0].get(
                        _qualified_name(_OFFICE_RELATIONSHIP_NAMESPACE, "embed")
                    )
                    linked_id = blips[0].get(
                        _qualified_name(_OFFICE_RELATIONSHIP_NAMESPACE, "link")
                    )
                    relationship = relationships.get(embed_id or "")
                    if (
                        embed_id is None
                        or linked_id is not None
                        or relationship is None
                        or relationship.relationship_type != _IMAGE_RELATIONSHIP
                        or relationship.is_external
                        or relationship.target not in archive_names
                    ):
                        has_unsupported = True
                        continue
                    used_relationship_ids.add(embed_id)
                    image_format = posixpath.splitext(relationship.target)[1].lstrip(".").lower()
                    image_data = archive.read(relationship.target)
                    drawings.append(
                        ImageSnapshot(
                            data=image_data,
                            image_format=image_format,
                            anchor=deepcopy(anchor),
                            anchor_coordinates=_anchor_coordinates(anchor),
                            has_supported_format=(
                                image_format in _SUPPORTED_IMAGE_FORMATS
                                and _has_supported_image_media(image_format, image_data)
                            ),
                            has_supported_anchor=type(anchor) in _SUPPORTED_ANCHOR_TYPES,
                        )
                    )
                    continue
                has_unsupported = True
            if set(relationships) != used_relationship_ids:
                has_unsupported = True
            profiles[name] = _DrawingPartProfile(tuple(drawings), has_unsupported)
    return profiles


def _sheet_drawing_state(
    sheet: Worksheet,
    profiles: dict[str, _DrawingPartProfile],
) -> tuple[tuple[DrawingSnapshot, ...], bool]:
    """Collect one worksheet's ordered drawings and unsupported state.

    Args:
        sheet: Loaded worksheet whose drawing relationships are inspected.
        profiles: Package drawing-part snapshots keyed by normalized path.

    Returns:
        Ordered supported drawings and whether any drawing content is unsupported.
    """

    drawings: list[DrawingSnapshot] = []
    has_unsupported = False
    for relationship in sheet._rels:
        if relationship.Type == _IMAGE_RELATIONSHIP:
            has_unsupported = True
            continue
        if relationship.Type != _DRAWING_RELATIONSHIP:
            continue
        target = relationship.Target.lstrip("/")
        profile = profiles.get(target)
        if profile is None:
            has_unsupported = True
            continue
        drawings.extend(profile.drawings)
        has_unsupported = has_unsupported or profile.has_unsupported_objects
    return tuple(drawings), has_unsupported


def _dimension_presentation(dimension: Any) -> DimensionPresentation:
    """Detach formatting shared by an openpyxl row or column dimension.

    Args:
        dimension: Openpyxl row or column dimension.

    Returns:
        Adapter-owned immutable presentation data.
    """

    return DimensionPresentation(
        hidden=bool(dimension.hidden),
        outline_level=int(dimension.outlineLevel or 0),
        collapsed=bool(dimension.collapsed),
        font=copy(dimension.font),
        fill=copy(dimension.fill),
        border=copy(dimension.border),
        alignment=copy(dimension.alignment),
        number_format=dimension.number_format,
        protection=copy(dimension.protection),
    )


def _read_rows(sheet: Worksheet) -> dict[int, RowPresentation]:
    """Snapshot explicitly configured worksheet row dimensions.

    Args:
        sheet: Openpyxl worksheet to read.

    Returns:
        Row numbers mapped to detached presentation data.
    """

    rows: dict[int, RowPresentation] = {}
    for index, dimension in sheet.row_dimensions.items():
        common = _dimension_presentation(dimension)
        rows[index] = RowPresentation(
            **vars(common),
            height=dimension.height,
            thick_top=bool(dimension.thickTop),
            thick_bottom=bool(dimension.thickBot),
        )
    return rows


def _read_columns(sheet: Worksheet) -> dict[int, ColumnPresentation]:
    """Expand and snapshot explicitly configured column dimensions.

    Args:
        sheet: Openpyxl worksheet to read.

    Returns:
        Column indexes mapped to detached presentation data.
    """

    columns: dict[int, ColumnPresentation] = {}
    for key, dimension in sheet.column_dimensions.items():
        common = _dimension_presentation(dimension)
        minimum = dimension.min or column_index_from_string(key)
        maximum = dimension.max or minimum
        for index in range(minimum, maximum + 1):
            columns[index] = ColumnPresentation(
                **vars(common),
                width=dimension.width,
                best_fit=bool(dimension.bestFit),
            )
    return columns


def _read_merges(sheet: Worksheet) -> tuple[Rectangle, ...]:
    """Convert openpyxl merged ranges into pure inclusive rectangles.

    Args:
        sheet: Openpyxl worksheet to inspect.

    Returns:
        Immutable merged rectangles in worksheet order.
    """

    ranges = cast(Iterable[CellRange], sheet.merged_cells.ranges)
    return tuple(
        Rectangle(
            cast(int, merged.min_row),
            cast(int, merged.min_col),
            cast(int, merged.max_row),
            cast(int, merged.max_col),
        )
        for merged in ranges
    )


def _optional_float(value: Any) -> float | None:
    """Convert an optional numeric openpyxl property to ``float``.

    Args:
        value: Optional numeric property.

    Returns:
        ``None`` or the equivalent float value.
    """

    return None if value is None else float(value)


def _merge_coordinates(merged_ranges: tuple[Rectangle, ...]) -> set[Coordinate]:
    """Enumerate every coordinate occupied by merged ranges.

    Args:
        merged_ranges: Inclusive merged rectangles.

    Returns:
        Coordinates belonging to at least one merge.
    """

    return {
        Coordinate(row, column)
        for merged in merged_ranges
        for row in range(merged.top, merged.bottom + 1)
        for column in range(merged.left, merged.right + 1)
    }


def _cell_presentation(cell: Any) -> CellPresentation:
    """Detach the writer-owned presentation of one worksheet cell.

    Args:
        cell: Openpyxl cell or merged-cell object.

    Returns:
        Immutable presentation snapshot used by the workbook writer.
    """

    return CellPresentation(
        font=copy(cell.font),
        fill=copy(cell.fill),
        border=copy(cell.border),
        alignment=copy(cell.alignment),
        number_format=cell.number_format,
        protection=copy(cell.protection),
        quote_prefix=bool(getattr(cell, "quotePrefix", False)),
        hyperlink=copy(getattr(cell, "hyperlink", None)),
        comment=copy(getattr(cell, "comment", None)),
        is_formula=getattr(cell, "data_type", None) == "f",
    )


def _read_sheet(
    sheet: Worksheet,
    drawing_profiles: dict[str, _DrawingPartProfile],
) -> SheetSnapshot:
    """Detach supported values, presentation, dimensions, and feature flags.

    Args:
        sheet: Openpyxl worksheet to snapshot.
        drawing_profiles: Package drawing-part classifications.

    Returns:
        Immutable adapter-owned worksheet state.
    """

    drawings, has_unsupported_drawings = _sheet_drawing_state(sheet, drawing_profiles)
    drawing_anchor_coordinates = {
        coordinate for drawing in drawings for coordinate in drawing.anchor_coordinates
    }
    merged_ranges = _read_merges(sheet)
    merge_coordinates = _merge_coordinates(merged_ranges)
    values: dict[Coordinate, Any] = {}
    presentations: dict[Coordinate, CellPresentation] = {}
    formula_cells: list[Coordinate] = []
    hyperlink_cells: list[Coordinate] = []
    comment_cells: list[Coordinate] = []

    for row in sheet.iter_rows():
        for cell in row:
            coordinate = Coordinate(cell.row, cell.column)
            is_material = (
                cell.value is not None
                or cell.has_style
                or getattr(cell, "hyperlink", None) is not None
                or getattr(cell, "comment", None) is not None
                or coordinate in merge_coordinates
            )
            if not is_material:
                continue
            value = cell.value if not isinstance(cell, MergedCell) else None
            is_formula = getattr(cell, "data_type", None) == "f"
            values[coordinate] = value
            presentations[coordinate] = _cell_presentation(cell)
            if is_formula:
                formula_cells.append(coordinate)
            if presentations[coordinate].hyperlink is not None:
                hyperlink_cells.append(coordinate)
            if presentations[coordinate].comment is not None:
                comment_cells.append(coordinate)

    synthetic_drawing_anchor_cells: set[Coordinate] = set()
    for coordinate in drawing_anchor_coordinates.difference(values):
        cell = sheet.cell(coordinate.row, coordinate.column)
        values[coordinate] = None
        presentations[coordinate] = _cell_presentation(cell)
        synthetic_drawing_anchor_cells.add(coordinate)

    freeze_panes = sheet.freeze_panes
    if freeze_panes is not None and not isinstance(freeze_panes, str):
        freeze_panes = freeze_panes.coordinate
    return SheetSnapshot(
        template=WorksheetTemplate(sheet.title, values, merged_ranges),
        cells=presentations,
        rows=_read_rows(sheet),
        columns=_read_columns(sheet),
        freeze_panes=freeze_panes,
        show_grid_lines=sheet.sheet_view.showGridLines,
        tab_color=copy(sheet.sheet_properties.tabColor),
        default_row_height=_optional_float(sheet.sheet_format.defaultRowHeight),
        default_column_width=_optional_float(sheet.sheet_format.defaultColWidth),
        conditional_formatting=copy(sheet.conditional_formatting),
        data_validations=copy(sheet.data_validations),
        auto_filter=copy(sheet.auto_filter),
        formula_cells=tuple(formula_cells),
        hyperlink_cells=tuple(hyperlink_cells),
        comment_cells=tuple(comment_cells),
        has_conditional_formatting=bool(len(sheet.conditional_formatting)),
        has_data_validations=bool(sheet.data_validations.count),
        has_tables=bool(sheet.tables),
        drawings=drawings,
        synthetic_drawing_anchor_cells=frozenset(synthetic_drawing_anchor_cells),
        has_unsupported_drawings=has_unsupported_drawings,
    )


def read_workbook(path: str | Path) -> WorkbookSnapshot:
    """Load a non-macro XLSX workbook and detach supported state.

    Args:
        path: Input ``.xlsx`` path.

    Returns:
        Immutable workbook snapshot containing no live openpyxl objects.
    """

    source_path = Path(path)
    drawing_profiles = _drawing_profiles(source_path)
    workbook = load_workbook(source_path, data_only=False, keep_links=False)
    try:
        return WorkbookSnapshot(
            sheets=tuple(_read_sheet(sheet, drawing_profiles) for sheet in workbook.worksheets),
            chartsheets=tuple(sheet.title for sheet in workbook.chartsheets),
            properties=copy(workbook.properties),
            loaded_theme=workbook.loaded_theme,
        )
    finally:
        workbook.close()
