from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SAMPLES = ROOT / "samples"
SAMPLE_STEMS = (
    "scalar_values",
    "repeated_blocks",
    "conditions_and_nesting",
    "cell_shift_lanes",
    "fixed_range_charts",
    "template_images",
    "regions",
    "polars_dataframe",
)


@pytest.mark.parametrize("stem", SAMPLE_STEMS)
def test_committed_sample_workbook_pairs_are_valid_and_rendered(stem: str) -> None:
    template_path = SAMPLES / f"{stem}_template.xlsx"
    output_path = SAMPLES / f"{stem}_output.xlsx"
    assert template_path.is_file()
    assert output_path.is_file()

    template = load_workbook(template_path, read_only=True, data_only=False)
    output = load_workbook(output_path, read_only=True, data_only=False)
    try:
        assert template.sheetnames == output.sheetnames
        assert any(
            "{%" in cell.value or "{{" in cell.value
            for sheet in template.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        )
        assert all(
            "{%" not in cell.value and "{{" not in cell.value
            for sheet in output.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        )
    finally:
        template.close()
        output.close()


def test_sample_catalog_names_every_committed_pair() -> None:
    catalog = (SAMPLES / "README.md").read_text(encoding="utf-8")

    for stem in SAMPLE_STEMS:
        assert f"{stem}_template.xlsx" in catalog
        assert f"{stem}_output.xlsx" in catalog
