from __future__ import annotations

from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from importlib import import_module
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excel_template_writer.diagnostics import ContextLocation, DiagnosticCode
from excel_template_writer.limits import ResourceLimits
from excel_template_writer.values import NormalizationResult, normalize_context
from excel_template_writer.xlsx import render_workbook

pl = pytest.importorskip("polars")
polars_adapters = import_module("excel_template_writer.adapters.polars").polars_adapters


def _codes_by_path(result: NormalizationResult) -> dict[str, DiagnosticCode]:
    return {
        diagnostic.location.path: diagnostic.code
        for diagnostic in result.diagnostics
        if isinstance(diagnostic.location, ContextLocation)
    }


def test_dataframe_becomes_ordered_canonical_records() -> None:
    frame = pl.DataFrame(
        {
            "name": ["Beta", "Alpha"],
            "amount": [Decimal("12.50"), Decimal("3.75")],
            "issued_on": [date(2026, 8, 19), date(2026, 8, 20)],
            "created_at": [datetime(2026, 8, 19, 9), datetime(2026, 8, 20, 10)],
            "optional": [None, "present"],
        },
        strict=False,
    )

    normalized = normalize_context(
        {"rows": frame},
        adapters=polars_adapters(),
    ).require()

    assert normalized["rows"] == (
        {
            "name": "Beta",
            "amount": Decimal("12.50"),
            "issued_on": date(2026, 8, 19),
            "created_at": datetime(2026, 8, 19, 9),
            "optional": None,
        },
        {
            "name": "Alpha",
            "amount": Decimal("3.75"),
            "issued_on": date(2026, 8, 20),
            "created_at": datetime(2026, 8, 20, 10),
            "optional": "present",
        },
    )


def test_null_and_nested_float_nan_become_canonical_null() -> None:
    frame = pl.DataFrame(
        {
            "metrics": [
                {"values": [1.0, float("nan")]},
                {"values": [None, 2.0]},
            ]
        }
    )

    normalized = normalize_context(
        {"rows": frame},
        adapters=polars_adapters(),
    ).require()

    assert normalized["rows"] == (
        {"metrics": {"values": (1.0, None)}},
        {"metrics": {"values": (None, 2.0)}},
    )


def test_empty_dataframe_becomes_an_empty_collection() -> None:
    frame = pl.DataFrame(schema={"name": pl.String, "amount": pl.Int64})

    normalized = normalize_context(
        {"rows": frame},
        adapters=polars_adapters(),
    ).require()

    assert normalized["rows"] == ()


def test_adapter_is_explicit_and_does_not_collect_lazy_frames() -> None:
    frame = pl.DataFrame({"name": ["Alpha"]})

    without_adapter = normalize_context({"rows": frame})
    lazy = normalize_context({"rows": frame.lazy()}, adapters=polars_adapters())

    assert _codes_by_path(without_adapter) == {
        "context.rows": DiagnosticCode.UNSUPPORTED_CONTEXT_VALUE
    }
    assert _codes_by_path(lazy) == {"context.rows": DiagnosticCode.UNSUPPORTED_CONTEXT_VALUE}


def test_canonical_validation_reports_unsupported_materialized_values() -> None:
    frame = pl.DataFrame(
        {
            "infinite": [float("inf")],
            "duration": [timedelta(days=1)],
            "binary": [b"opaque"],
        }
    )

    result = normalize_context({"rows": frame}, adapters=polars_adapters())

    assert _codes_by_path(result) == {
        "context.rows[0].infinite": DiagnosticCode.NON_FINITE_CONTEXT_NUMBER,
        "context.rows[0].duration": DiagnosticCode.UNSUPPORTED_CONTEXT_VALUE,
        "context.rows[0].binary": DiagnosticCode.UNSUPPORTED_CONTEXT_VALUE,
    }


@pytest.mark.parametrize(
    "series",
    [
        pl.Series("when", [datetime(2026, 8, 19, 9)], dtype=pl.Datetime("ns")),
        pl.Series("when", [time(9, 30)], dtype=pl.Time),
        pl.Series(
            "when",
            [datetime(2026, 8, 19, 9, tzinfo=UTC)],
            dtype=pl.Datetime("us", "UTC"),
        ),
    ],
    ids=["nanosecond-datetime", "time", "timezone-aware-datetime"],
)
def test_temporal_values_that_cannot_be_preserved_are_rejected_before_conversion(
    series: object,
) -> None:
    frame = pl.DataFrame([series])

    result = normalize_context({"rows": frame}, adapters=polars_adapters())

    assert _codes_by_path(result) == {"context.rows": DiagnosticCode.VALUE_ADAPTER_FAILED}


def test_resource_limits_apply_after_dataframe_conversion() -> None:
    frame = pl.DataFrame({"name": ["Alpha", "Beta"]})
    limits = ResourceLimits(max_container_items=1)

    result = normalize_context(
        {"rows": frame},
        adapters=polars_adapters(),
        limits=limits,
    )

    assert _codes_by_path(result) == {
        "context.rows": DiagnosticCode.CONTEXT_RESOURCE_LIMIT_EXCEEDED
    }


def test_dataframe_renders_through_the_xlsx_entrypoint(tmp_path: Path) -> None:
    template_path = tmp_path / "polars-template.xlsx"
    output_path = tmp_path / "polars-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for row in rows %}{{ row.name }}"
    sheet["B1"] = "{{ row.amount }}{% endfor %}"
    workbook.save(template_path)
    workbook.close()
    frame = pl.DataFrame({"name": ["Alpha", "Beta"], "amount": [10, 20]})

    render_workbook(
        template_path,
        output_path,
        {"rows": frame},
        adapters=polars_adapters(),
    )

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        assert sheet["A1"].value == "Alpha"
        assert sheet["B1"].value == 10
        assert sheet["A2"].value == "Beta"
        assert sheet["B2"].value == 20
    finally:
        rendered.close()
