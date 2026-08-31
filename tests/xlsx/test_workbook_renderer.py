from __future__ import annotations

import base64
from dataclasses import replace
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, TwoCellAnchor
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

from excel_template_writer.diagnostics import (
    DiagnosticCode,
    TemplateCompilationError,
    TemplateRenderError,
)
from excel_template_writer.limits import ResourceLimits
from excel_template_writer.values import TypeAdapter
from excel_template_writer.xlsx import render_workbook


class _StaticImage(Image):
    def __init__(self, image_format: str = "png", data: bytes | None = None) -> None:
        """Create a Pillow-independent image for unsupported-drawing tests."""

        self.width = 1
        self.height = 1
        self.format = image_format
        self._payload = data or base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUB"
            "AScY42YAAAAASUVORK5CYII="
        )

    def _data(self) -> bytes:
        """Return the configured embedded image payload."""

        return self._payload


_ONE_PIXEL_JPEG = base64.b64decode(
    "/9j/4AAQSkZJRgABAQEAYABgAAD/2wBDAAMCAgMCAgMDAwMEAwMEBQgFBQQEBQoHBwYIDAoM"
    "DAsKCwsNDhIQDQ4RDgsLEBYQERMUFRUVDA8XGBYUGBIUFRT/2wBDAQMEBAUEBQkFBQkUDQsN"
    "FBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBT/wAA"
    "RCAABAAEDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAA"
    "AgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkK"
    "FhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWG"
    "h4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl"
    "5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREA"
    "AgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYk"
    "NOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOE"
    "hYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk"
    "5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD896KKK/1TPhz/2Q=="
)


def _save(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    return path


def _uncompressed_size(path: Path) -> int:
    with ZipFile(path) as archive:
        return sum(member.file_size for member in archive.infolist())


def _chart_references(chart: Any) -> set[str]:
    tree = chart._write()
    return {
        element.text
        for element in tree.iter()
        if element.tag.rsplit("}", 1)[-1] == "f" and element.text
    }


def _chart_style(path: Path, part: str = "chart1.xml") -> str | None:
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(f"xl/charts/{part}"))
    style = next(
        (element for element in root.iter() if element.tag.rsplit("}", 1)[-1] == "style"),
        None,
    )
    return None if style is None else style.get("val")


_DRAWING_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DRAWING_MAIN_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/main"
_CHART_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_PACKAGE_RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_TYPES_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/content-types"
_DRAWING_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
)


def _qn(namespace: str, name: str) -> str:
    return f"{{{namespace}}}{name}"


def _shape_anchor(
    *,
    row: int,
    column: int,
    text: str = "Static {{ customer.name }}",
    preset: str = "roundRect",
    to_row: int | None = None,
    to_column: int | None = None,
    unsupported: str | None = None,
) -> ElementTree.Element:
    if to_row is None or to_column is None:
        anchor = ElementTree.Element(_qn(_DRAWING_NAMESPACE, "oneCellAnchor"))
    else:
        anchor = ElementTree.Element(_qn(_DRAWING_NAMESPACE, "twoCellAnchor"))

    def add_marker(name: str, marker_row: int, marker_column: int) -> None:
        marker = ElementTree.SubElement(anchor, _qn(_DRAWING_NAMESPACE, name))
        ElementTree.SubElement(marker, _qn(_DRAWING_NAMESPACE, "col")).text = str(marker_column)
        ElementTree.SubElement(marker, _qn(_DRAWING_NAMESPACE, "colOff")).text = "9525"
        ElementTree.SubElement(marker, _qn(_DRAWING_NAMESPACE, "row")).text = str(marker_row)
        ElementTree.SubElement(marker, _qn(_DRAWING_NAMESPACE, "rowOff")).text = "19050"

    add_marker("from", row, column)
    if to_row is None or to_column is None:
        ElementTree.SubElement(
            anchor,
            _qn(_DRAWING_NAMESPACE, "ext"),
            {"cx": "2286000", "cy": "762000"},
        )
    else:
        add_marker("to", to_row, to_column)

    shape = ElementTree.SubElement(anchor, _qn(_DRAWING_NAMESPACE, "sp"))
    if unsupported == "textlink":
        shape.set("textlink", "Sheet!A1")
    if unsupported == "macro":
        shape.set("macro", "ExampleMacro")
    non_visual = ElementTree.SubElement(shape, _qn(_DRAWING_NAMESPACE, "nvSpPr"))
    ElementTree.SubElement(
        non_visual,
        _qn(_DRAWING_NAMESPACE, "cNvPr"),
        {"id": "7", "name": "Decorative text", "descr": "Editable callout"},
    )
    ElementTree.SubElement(
        non_visual,
        _qn(_DRAWING_NAMESPACE, "cNvSpPr"),
        {"txBox": "1"},
    )
    shape_properties = ElementTree.SubElement(shape, _qn(_DRAWING_NAMESPACE, "spPr"))
    transform = ElementTree.SubElement(
        shape_properties,
        _qn(_DRAWING_MAIN_NAMESPACE, "xfrm"),
        {"rot": "600000"},
    )
    ElementTree.SubElement(
        transform,
        _qn(_DRAWING_MAIN_NAMESPACE, "off"),
        {"x": "0", "y": "0"},
    )
    ElementTree.SubElement(
        transform,
        _qn(_DRAWING_MAIN_NAMESPACE, "ext"),
        {"cx": "2286000", "cy": "762000"},
    )
    geometry = ElementTree.SubElement(
        shape_properties,
        _qn(_DRAWING_MAIN_NAMESPACE, "prstGeom"),
        {"prst": preset},
    )
    ElementTree.SubElement(geometry, _qn(_DRAWING_MAIN_NAMESPACE, "avLst"))
    fill = ElementTree.SubElement(shape_properties, _qn(_DRAWING_MAIN_NAMESPACE, "solidFill"))
    ElementTree.SubElement(
        fill,
        _qn(_DRAWING_MAIN_NAMESPACE, "srgbClr"),
        {"val": "D9EAF7"},
    )
    line = ElementTree.SubElement(
        shape_properties,
        _qn(_DRAWING_MAIN_NAMESPACE, "ln"),
        {"w": "19050"},
    )
    line_fill = ElementTree.SubElement(line, _qn(_DRAWING_MAIN_NAMESPACE, "solidFill"))
    ElementTree.SubElement(
        line_fill,
        _qn(_DRAWING_MAIN_NAMESPACE, "srgbClr"),
        {"val": "1F4E78"},
    )
    text_body = ElementTree.SubElement(shape, _qn(_DRAWING_NAMESPACE, "txBody"))
    body_properties = ElementTree.SubElement(
        text_body,
        _qn(_DRAWING_MAIN_NAMESPACE, "bodyPr"),
        {"wrap": "square", "anchor": "ctr", "lIns": "91440", "rIns": "91440"},
    )
    if unsupported == "wordart":
        ElementTree.SubElement(
            body_properties,
            _qn(_DRAWING_MAIN_NAMESPACE, "prstTxWarp"),
            {"prst": "textArchUp"},
        )
    else:
        ElementTree.SubElement(body_properties, _qn(_DRAWING_MAIN_NAMESPACE, "spAutoFit"))
    ElementTree.SubElement(text_body, _qn(_DRAWING_MAIN_NAMESPACE, "lstStyle"))
    paragraph = ElementTree.SubElement(text_body, _qn(_DRAWING_MAIN_NAMESPACE, "p"))
    if unsupported == "field":
        field = ElementTree.SubElement(
            paragraph,
            _qn(_DRAWING_MAIN_NAMESPACE, "fld"),
            {"id": "{00000000-0000-0000-0000-000000000001}", "type": "datetime"},
        )
        ElementTree.SubElement(field, _qn(_DRAWING_MAIN_NAMESPACE, "rPr"))
        ElementTree.SubElement(field, _qn(_DRAWING_MAIN_NAMESPACE, "t")).text = text
    else:
        run = ElementTree.SubElement(paragraph, _qn(_DRAWING_MAIN_NAMESPACE, "r"))
        run_properties = ElementTree.SubElement(
            run,
            _qn(_DRAWING_MAIN_NAMESPACE, "rPr"),
            {"lang": "en-US", "sz": "1400", "b": "1"},
        )
        run_fill = ElementTree.SubElement(
            run_properties,
            _qn(_DRAWING_MAIN_NAMESPACE, "solidFill"),
        )
        ElementTree.SubElement(
            run_fill,
            _qn(_DRAWING_MAIN_NAMESPACE, "srgbClr"),
            {"val": "1F1F1F"},
        )
        ElementTree.SubElement(run, _qn(_DRAWING_MAIN_NAMESPACE, "t")).text = text
    ElementTree.SubElement(paragraph, _qn(_DRAWING_MAIN_NAMESPACE, "endParaRPr"))
    ElementTree.SubElement(
        anchor,
        _qn(_DRAWING_NAMESPACE, "clientData"),
        {"fLocksWithSheet": "1", "fPrintsWithSheet": "1"},
    )
    return anchor


def _inject_text_shapes(path: Path, anchors: tuple[ElementTree.Element, ...]) -> None:
    temporary_path = path.with_name(f"{path.stem}-with-shapes.xlsx")
    with ZipFile(path, "r") as source, ZipFile(temporary_path, "w") as destination:
        source_names = frozenset(source.namelist())
        replacements: dict[str, bytes] = {}
        additions: dict[str, bytes] = {}
        drawing_part = "xl/drawings/drawing1.xml"
        if drawing_part in source_names:
            drawing_root = ElementTree.fromstring(source.read(drawing_part))
            drawing_root.extend(anchors)
            replacements[drawing_part] = ElementTree.tostring(drawing_root, encoding="utf-8")
        else:
            drawing_root = ElementTree.Element(_qn(_DRAWING_NAMESPACE, "wsDr"))
            drawing_root.extend(anchors)
            additions[drawing_part] = ElementTree.tostring(drawing_root, encoding="utf-8")

            relationships_part = "xl/worksheets/_rels/sheet1.xml.rels"
            if relationships_part in source_names:
                relationships_root = ElementTree.fromstring(source.read(relationships_part))
            else:
                relationships_root = ElementTree.Element(
                    _qn(_PACKAGE_RELATIONSHIP_NAMESPACE, "Relationships")
                )
            relationship_id = "rId1"
            used_ids = {relationship.get("Id") for relationship in relationships_root}
            while relationship_id in used_ids:
                relationship_id = f"rId{int(relationship_id.removeprefix('rId')) + 1}"
            ElementTree.SubElement(
                relationships_root,
                _qn(_PACKAGE_RELATIONSHIP_NAMESPACE, "Relationship"),
                {
                    "Id": relationship_id,
                    "Type": _DRAWING_RELATIONSHIP,
                    "Target": "/xl/drawings/drawing1.xml",
                },
            )
            relationship_xml = ElementTree.tostring(relationships_root, encoding="utf-8")
            if relationships_part in source_names:
                replacements[relationships_part] = relationship_xml
            else:
                additions[relationships_part] = relationship_xml

            worksheet_part = "xl/worksheets/sheet1.xml"
            worksheet_root = ElementTree.fromstring(source.read(worksheet_part))
            ElementTree.SubElement(
                worksheet_root,
                _qn(_SPREADSHEET_NAMESPACE, "drawing"),
                {_qn(_OFFICE_RELATIONSHIP_NAMESPACE, "id"): relationship_id},
            )
            replacements[worksheet_part] = ElementTree.tostring(
                worksheet_root,
                encoding="utf-8",
            )

            content_types = ElementTree.fromstring(source.read("[Content_Types].xml"))
            ElementTree.SubElement(
                content_types,
                _qn(_CONTENT_TYPES_NAMESPACE, "Override"),
                {
                    "PartName": "/xl/drawings/drawing1.xml",
                    "ContentType": ("application/vnd.openxmlformats-officedocument.drawing+xml"),
                },
            )
            replacements["[Content_Types].xml"] = ElementTree.tostring(
                content_types,
                encoding="utf-8",
            )

        for member in source.infolist():
            destination.writestr(
                member,
                replacements.get(member.filename, source.read(member.filename)),
            )
        for name, data in additions.items():
            destination.writestr(name, data)
    temporary_path.replace(path)


def _drawing_anchors(path: Path, part: str = "drawing1.xml") -> list[ElementTree.Element]:
    with ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(f"xl/drawings/{part}"))
    return list(root)


def _worksheet_drawing_target(path: Path, sheet_index: int = 1) -> str:
    worksheet_part = f"xl/worksheets/sheet{sheet_index}.xml"
    relationships_part = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
    with ZipFile(path) as archive:
        worksheet = ElementTree.fromstring(archive.read(worksheet_part))
        root = ElementTree.fromstring(archive.read(relationships_part))
    drawing = worksheet.find(_qn(_SPREADSHEET_NAMESPACE, "drawing"))
    assert drawing is not None
    relationship_id = drawing.get(_qn(_OFFICE_RELATIONSHIP_NAMESPACE, "id"))
    relationships = [
        relationship
        for relationship in root
        if relationship.get("Id") == relationship_id
        and relationship.get("Type") == _DRAWING_RELATIONSHIP
    ]
    assert len(relationships) == 1
    target = relationships[0].get("Target")
    assert target is not None
    return target.lstrip("/")


def _drawing_kind(anchor: ElementTree.Element) -> str:
    if anchor.find(f".//{{{_CHART_NAMESPACE}}}chart") is not None:
        return "chart"
    if anchor.find(f".//{{{_DRAWING_MAIN_NAMESPACE}}}blip") is not None:
        return "image"
    if anchor.find(f".//{{{_DRAWING_NAMESPACE}}}sp") is not None:
        return "text shape"
    return "unsupported"


def _anchor_marker(anchor: ElementTree.Element, marker: str = "from") -> tuple[int, int]:
    element = anchor.find(f"{{{_DRAWING_NAMESPACE}}}{marker}")
    assert element is not None
    row = element.find(f"{{{_DRAWING_NAMESPACE}}}row")
    column = element.find(f"{{{_DRAWING_NAMESPACE}}}col")
    assert row is not None and row.text is not None
    assert column is not None and column.text is not None
    return int(row.text), int(column.text)


def _drawing_ext(anchor: ElementTree.Element) -> tuple[str | None, str | None]:
    extent = anchor.find(f"{{{_DRAWING_NAMESPACE}}}ext")
    assert extent is not None
    return extent.get("cx"), extent.get("cy")


def _picture_metadata(anchor: ElementTree.Element) -> tuple[str | None, str | None]:
    properties = anchor.find(f".//{{{_DRAWING_NAMESPACE}}}cNvPr")
    assert properties is not None
    return properties.get("name"), properties.get("descr")


def _shape_text(anchor: ElementTree.Element) -> str:
    return "".join(element.text or "" for element in anchor.iter(_qn(_DRAWING_MAIN_NAMESPACE, "t")))


def _shape_visual_state(anchor: ElementTree.Element) -> tuple[object, ...]:
    geometry = anchor.find(f".//{{{_DRAWING_MAIN_NAMESPACE}}}prstGeom")
    transform = anchor.find(f".//{{{_DRAWING_MAIN_NAMESPACE}}}xfrm")
    body = anchor.find(f".//{{{_DRAWING_MAIN_NAMESPACE}}}bodyPr")
    run = anchor.find(f".//{{{_DRAWING_MAIN_NAMESPACE}}}rPr")
    metadata = anchor.find(f".//{{{_DRAWING_NAMESPACE}}}cNvPr")
    assert geometry is not None
    assert transform is not None
    assert body is not None
    assert run is not None
    assert metadata is not None
    return (
        geometry.get("prst"),
        transform.get("rot"),
        body.get("wrap"),
        body.get("anchor"),
        body.get("lIns"),
        run.get("sz"),
        run.get("b"),
        metadata.get("name"),
        metadata.get("descr"),
    )


def _media_payloads(path: Path) -> list[bytes]:
    with ZipFile(path) as archive:
        names = sorted(name for name in archive.namelist() if name.startswith("xl/media/"))
        return [archive.read(name) for name in names]


def _set_drawing_order(path: Path, order: tuple[str, ...]) -> None:
    temporary_path = path.with_name(f"{path.stem}-reordered.xlsx")
    with ZipFile(path, "r") as source, ZipFile(temporary_path, "w") as destination:
        for member in source.infolist():
            data = source.read(member.filename)
            if member.filename == "xl/drawings/drawing1.xml":
                root = ElementTree.fromstring(data)
                anchors = list(root)
                selected: list[ElementTree.Element] = []
                for kind in order:
                    selected.append(
                        next(anchor for anchor in anchors if _drawing_kind(anchor) == kind)
                    )
                root[:] = selected
                data = ElementTree.tostring(root, encoding="utf-8")
            destination.writestr(member, data)
    temporary_path.replace(path)


def test_render_workbook_normalizes_adapter_values_once_for_all_sheets(
    tmp_path: Path,
) -> None:
    class Rows:
        def __init__(self) -> None:
            self.values = [{"name": "Alpha"}, {"name": "Beta"}]

    template_path = tmp_path / "adapter-template.xlsx"
    output_path = tmp_path / "adapter-output.xlsx"
    workbook = Workbook()
    workbook.active.title = "First"
    workbook.active["A1"] = "{% for row in rows %}{{ row.name }}{% endfor %}"
    second = workbook.create_sheet("Second")
    second["A1"] = "{% for row in rows %}{{ row.name }}{% endfor %}"
    _save(workbook, template_path)
    calls = 0

    def convert(value: Rows) -> list[dict[str, str]]:
        nonlocal calls
        calls += 1
        return value.values

    render_workbook(
        template_path,
        output_path,
        {"rows": Rows()},
        adapters=(TypeAdapter(Rows, convert),),
    )

    assert calls == 1
    rendered = load_workbook(output_path)
    try:
        for sheet in rendered.worksheets:
            assert sheet["A1"].value == "Alpha"
            assert sheet["A2"].value == "Beta"
    finally:
        rendered.close()


def test_render_workbook_copies_values_direct_styles_styled_blanks_and_dimensions(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["B"].hidden = True
    sheet.row_dimensions[2].height = 31
    sheet.row_dimensions[2].outlineLevel = 1
    sheet["A1"] = "Description"
    sheet["B1"] = "Amount"
    sheet["A2"] = "{% for line in lines %}{{ line.description }}"
    sheet["B2"] = "{{ line.amount }}{% endfor %}"
    sheet["A3"] = None
    sheet["B3"] = "Footer"

    thin = Side(style="thin", color="FF445566")
    for cell in sheet[2]:
        cell.font = Font(name="Aptos", bold=True, color="FF17365D")
        cell.fill = PatternFill("solid", fgColor="FFEAF3F8")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.protection = Protection(locked=False, hidden=True)
    sheet["B2"].number_format = "$#,##0.00;[Red]-$#,##0.00"

    # This cell has no value but is materially part of the workbook presentation.
    sheet["A3"].fill = PatternFill("solid", fgColor="FFFFF2CC")
    sheet["A3"].border = Border(bottom=thin)
    _save(workbook, template_path)

    result = render_workbook(
        template_path,
        output_path,
        {
            "lines": [
                {"description": "Consulting", "amount": 1250},
                {"description": "Support", "amount": 350},
            ]
        },
    )

    assert result.output_path == output_path
    assert result.diagnostics == ()
    template = load_workbook(template_path)
    rendered = load_workbook(output_path)
    try:
        source = template["Report"]
        target = rendered["Report"]
        assert source["A2"].value.startswith("{% for")
        assert target["A2"].value == "Consulting"
        assert target["A3"].value == "Support"
        assert target["B2"].value == 1250
        assert target["B3"].value == 350
        assert target["B4"].value == "Footer"
        assert target["A4"].value is None
        assert target["A4"].fill.fgColor.rgb == "FFFFF2CC"
        assert target["A4"].border.bottom.style == "thin"

        for coordinate in ("A2", "A3"):
            cell = target[coordinate]
            assert cell.font.bold is True
            assert cell.font.color.rgb == "FF17365D"
            assert cell.fill.fgColor.rgb == "FFEAF3F8"
            assert cell.border.left.style == "thin"
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.wrap_text is True
            assert cell.protection.locked is False
            assert cell.protection.hidden is True
        assert target["B2"].number_format == "$#,##0.00;[Red]-$#,##0.00"
        assert target["B3"].number_format == "$#,##0.00;[Red]-$#,##0.00"
        assert target.row_dimensions[2].height == 31
        assert target.row_dimensions[3].height == 31
        assert target.row_dimensions[2].outlineLevel == 1
        assert target.row_dimensions[3].outlineLevel == 1
        assert target.column_dimensions["A"].width == 26
        assert target.column_dimensions["B"].width == 18
        assert target.column_dimensions["B"].hidden is True
    finally:
        template.close()
        rendered.close()


def test_render_workbook_repeats_contained_merges_and_row_heights(tmp_path: Path) -> None:
    template_path = tmp_path / "cards.xlsx"
    output_path = tmp_path / "cards-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cards"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "{% for card in cards %}{{ card.title }}"
    sheet["A2"] = "Owner"
    sheet["B2"] = "{{ card.owner }}"
    sheet["C2"] = "{{ card.amount }}{% endfor %}"
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 19
    sheet["A3"] = "After cards"
    sheet["A1"].fill = PatternFill("solid", fgColor="FF17365D")
    sheet["A1"].font = Font(bold=True, color="FFFFFFFF")
    sheet["C2"].number_format = "$#,##0"
    _save(workbook, template_path)

    render_workbook(
        template_path,
        output_path,
        {
            "cards": [
                {"title": "Alpha", "owner": "Mina", "amount": 100},
                {"title": "Beta", "owner": "Jules", "amount": 200},
            ]
        },
    )

    rendered = load_workbook(output_path)
    try:
        target = rendered["Cards"]
        assert {str(item) for item in target.merged_cells.ranges} == {"A1:C1", "A3:C3"}
        assert target["A1"].value == "Alpha"
        assert target["A3"].value == "Beta"
        assert target["A5"].value == "After cards"
        assert [target.row_dimensions[row].height for row in range(1, 5)] == [28, 19, 28, 19]
        assert target["A3"].fill.fgColor.rgb == "FF17365D"
        assert target["C4"].number_format == "$#,##0"
    finally:
        rendered.close()


def test_render_workbook_preserves_native_value_types(tmp_path: Path) -> None:
    template_path = tmp_path / "types.xlsx"
    output_path = tmp_path / "types-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{{ issued_on }}"
    sheet["A1"].number_format = "yyyy-mm-dd"
    sheet["B1"] = "{{ approved }}"
    sheet["C1"] = "{{ total }}"
    sheet["D1"] = "{{ exact_amount }}"
    sheet["D1"].number_format = "$#,##0.00"
    sheet["E1"] = "{{ cutoff }}"
    sheet["E1"].number_format = "hh:mm:ss"
    _save(workbook, template_path)

    render_workbook(
        template_path,
        output_path,
        {
            "issued_on": date(2026, 8, 13),
            "approved": True,
            "total": 42.5,
            "exact_amount": Decimal("19.75"),
            "cutoff": time(17, 30),
        },
    )

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        issued_on = sheet["A1"].value
        assert isinstance(issued_on, (date, datetime))
        assert sheet["A1"].data_type == "d"
        assert sheet["B1"].value is True
        assert sheet["B1"].data_type == "b"
        assert sheet["C1"].value == 42.5
        assert sheet["C1"].data_type == "n"
        assert sheet["D1"].value == 19.75
        assert sheet["D1"].data_type == "n"
        assert sheet["D1"].number_format == "$#,##0.00"
        assert sheet["E1"].value == time(17, 30)
        assert sheet["E1"].data_type == "d"
        assert sheet["E1"].number_format == "hh:mm:ss"
    finally:
        rendered.close()


def test_date_filter_writes_text_while_unfiltered_date_remains_native(tmp_path: Path) -> None:
    template_path = tmp_path / "date-filter.xlsx"
    output_path = tmp_path / "date-filter-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{{ report_date }}"
    sheet["A1"].number_format = '"For the month ending "dd mmmm yyyy'
    sheet["B1"] = '{{ report_date | date("YYYY-mm") }}'
    sheet["B1"].number_format = "yyyy-mm-dd"
    sheet["C1"] = 'For the month ending {{ report_date | date("dd mmmm yyyy") }}'
    _save(workbook, template_path)

    render_workbook(
        template_path,
        output_path,
        {"report_date": date(2026, 8, 31)},
    )

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        assert isinstance(sheet["A1"].value, (date, datetime))
        assert sheet["A1"].data_type == "d"
        assert sheet["A1"].number_format == '"For the month ending "dd mmmm yyyy'
        assert sheet["B1"].value == "2026-08"
        assert sheet["B1"].data_type == "s"
        assert sheet["B1"].number_format == "yyyy-mm-dd"
        assert sheet["C1"].value == "For the month ending 31 August 2026"
        assert sheet["C1"].data_type == "s"
    finally:
        rendered.close()


def test_aggregate_filters_and_arithmetic_write_typed_values_with_formatting(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "aggregate-arithmetic.xlsx"
    output_path = tmp_path / "aggregate-arithmetic-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{{ amounts | sum }}"
    sheet["B1"] = '{{ rows | sum("amount") }}'
    sheet["B1"].number_format = "$#,##0.00"
    sheet["C1"] = '{{ rows | min("amount") }}'
    sheet["D1"] = '{{ rows | max("amount") }}'
    sheet["E1"] = "{{ rows | count }}"
    sheet["F1"] = '{{ (rows | max("amount")) - (rows | min("amount")) }}'
    sheet["F1"].number_format = "$#,##0.00"
    sheet["G1"] = '{{ empty_rows | max("amount") }}'
    sheet["H1"] = "{{ ((base + adjustment) * factor - discount) / divisor }}"
    sheet["H1"].number_format = "0.00"
    _save(workbook, template_path)

    render_workbook(
        template_path,
        output_path,
        {
            "amounts": [1, None, 2.5],
            "rows": [{"amount": Decimal("12.50")}, {"amount": 2}, {"amount": None}],
            "empty_rows": [],
            "base": 100,
            "adjustment": 20,
            "factor": 2,
            "discount": 40,
            "divisor": 4,
        },
    )

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        assert sheet["A1"].value == 3.5
        assert sheet["A1"].data_type == "n"
        assert sheet["B1"].value == 14.5
        assert sheet["B1"].data_type == "n"
        assert sheet["B1"].number_format == "$#,##0.00"
        assert sheet["C1"].value == 2
        assert sheet["C1"].data_type == "n"
        assert sheet["D1"].value == 12.5
        assert sheet["D1"].data_type == "n"
        assert sheet["E1"].value == 3
        assert sheet["E1"].data_type == "n"
        assert sheet["F1"].value == 10.5
        assert sheet["F1"].data_type == "n"
        assert sheet["F1"].number_format == "$#,##0.00"
        assert sheet["G1"].value is None
        assert sheet["H1"].value == 50
        assert sheet["H1"].data_type == "n"
        assert sheet["H1"].number_format == "0.00"
    finally:
        rendered.close()


def test_repeats_blank_cells_that_carry_formatting(tmp_path: Path) -> None:
    template_path = tmp_path / "styled-blanks.xlsx"
    output_path = tmp_path / "styled-blanks-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = None
    sheet["B1"].fill = PatternFill("solid", fgColor="FFFFF2CC")
    sheet["B1"].border = Border(bottom=Side(style="double", color="FF000000"))
    sheet["C1"] = "{% endfor %}"
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": ["A", "B"]})

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        for coordinate in ("B1", "B2"):
            assert sheet[coordinate].value is None
            assert sheet[coordinate].fill.fgColor.rgb == "FFFFF2CC"
            assert sheet[coordinate].border.bottom.style == "double"
    finally:
        rendered.close()


def test_cell_shift_region_moves_its_full_formatted_band_and_keeps_adjacent_cells(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "region-template.xlsx"
    output_path = tmp_path / "region-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Region"
    sheet["A1"] = '{% region shift="cells" %}'
    sheet["D2"] = '{% for item in items shift="cells" %}{{ item }}'
    sheet["F2"] = "{% endfor %}"
    sheet["J2"] = "{% endregion %}"
    sheet["A10"] = "Moves with A:J"
    sheet["A10"].fill = PatternFill("solid", fgColor="FFBDD7EE")
    sheet["J10"].fill = PatternFill("solid", fgColor="FFBDD7EE")
    sheet["K10"] = "Stays in K:P"
    sheet["K10"].fill = PatternFill("solid", fgColor="FFFFE699")
    sheet["D2"].font = Font(bold=True, color="FF17365D")
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": ["One", "Two", "Three", "Four", "Five"]})

    rendered = load_workbook(output_path)
    try:
        sheet = rendered["Region"]
        assert [sheet[f"D{row}"].value for row in range(2, 7)] == [
            "One",
            "Two",
            "Three",
            "Four",
            "Five",
        ]
        assert all(sheet[f"D{row}"].font.bold for row in range(2, 7))
        assert sheet["A14"].value == "Moves with A:J"
        assert sheet["A14"].fill.fgColor.rgb == "FFBDD7EE"
        assert sheet["J14"].value is None
        assert sheet["J14"].fill.fgColor.rgb == "FFBDD7EE"
        assert sheet["K10"].value == "Stays in K:P"
        assert sheet["K10"].fill.fgColor.rgb == "FFFFE699"
        assert sheet["A10"].value is None
    finally:
        rendered.close()


@pytest.mark.parametrize("feature", ["formula", "conditional-formatting"])
def test_rejects_affected_unsupported_features(tmp_path: Path, feature: str) -> None:
    template_path = tmp_path / f"{feature}.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}{% endfor %}"
    if feature == "formula":
        sheet["B2"] = "=SUM(A1:A1)"
        expected = DiagnosticCode.FORMULA_REQUIRES_UNSUPPORTED_TRANSFORM
    else:
        sheet.conditional_formatting.add(
            "A1:A3",
            CellIsRule(operator="greaterThan", formula=["0"]),
        )
        expected = DiagnosticCode.CONDITIONAL_FORMATTING_REQUIRES_UNSUPPORTED_TRANSFORM
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert expected in {diagnostic.code for diagnostic in caught.value.diagnostics}
    assert not output_path.exists()


def test_rejects_merge_that_crosses_a_repeat_boundary(tmp_path: Path) -> None:
    template_path = tmp_path / "crossing-merge.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}{% endfor %}"
    sheet.merge_cells("A1:B1")
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert DiagnosticCode.MERGE_CROSSES_BLOCK_BOUNDARY in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_moves_static_merge_below_row_expansion(tmp_path: Path) -> None:
    template_path = tmp_path / "moving-merge.xlsx"
    output_path = tmp_path / "moving-merge-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}{% endfor %}"
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Merged footer"
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": ["A", "B"]})

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        assert {str(item) for item in sheet.merged_cells.ranges} == {"A4:B4"}
        assert sheet["A4"].value == "Merged footer"
    finally:
        rendered.close()


def test_preserves_unaffected_formula_and_conditional_formatting(tmp_path: Path) -> None:
    template_path = tmp_path / "static-features.xlsx"
    output_path = tmp_path / "static-features-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "=1+1"
    sheet["B1"] = "{{ label }}"
    sheet.conditional_formatting.add(
        "A1:A2",
        CellIsRule(operator="greaterThan", formula=["0"]),
    )
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"label": "Static layout"})

    rendered = load_workbook(output_path, data_only=False)
    try:
        sheet = rendered.active
        assert sheet["A1"].value == "=1+1"
        assert sheet["A1"].data_type == "f"
        assert sheet["B1"].value == "Static layout"
        assert len(sheet.conditional_formatting) == 1
    finally:
        rendered.close()


def test_preserves_fixed_chart_references_when_a_cell_shift_repeat_exceeds_them(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "fixed-chart-template.xlsx"
    output_path = tmp_path / "fixed-chart-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet["A1"] = "Identifier"
    sheet["B1"] = "Value"
    sheet["A2"] = '{% for item in items shift="cells" %}{{ item.identifier }}'
    sheet["B2"] = "{{ item.value }}{% endfor %}"
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "First nine items"
    chart.display_blanks = "gap"
    chart.add_data(
        Reference(sheet, min_col=2, min_row=1, max_row=10),
        titles_from_data=True,
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=10))
    sheet.add_chart(chart, "D2")
    _save(workbook, template_path)

    render_workbook(
        template_path,
        output_path,
        {"items": [{"identifier": f"Item {index}", "value": index * 10} for index in range(1, 13)]},
    )

    rendered = load_workbook(output_path)
    try:
        sheet = rendered["Sales"]
        assert [sheet[f"A{row}"].value for row in range(2, 14)] == [
            f"Item {index}" for index in range(1, 13)
        ]
        assert len(sheet._charts) == 1
        rendered_chart = sheet._charts[0]
        assert type(rendered_chart) is BarChart
        assert rendered_chart.display_blanks == "gap"
        assert isinstance(rendered_chart.anchor, OneCellAnchor)
        assert (rendered_chart.anchor._from.row, rendered_chart.anchor._from.col) == (1, 3)
        assert _chart_references(rendered_chart) == {
            "'Sales'!B1",
            "'Sales'!$A$2:$A$10",
            "'Sales'!$B$2:$B$10",
        }
        assert _chart_style(template_path) == "10"
        assert _chart_style(output_path) == "10"
    finally:
        rendered.close()


def test_moves_chart_anchor_down_with_row_expansion(tmp_path: Path) -> None:
    template_path = tmp_path / "moving-chart-template.xlsx"
    output_path = tmp_path / "moving-chart-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    sheet.add_chart(chart, "D3")
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": [1, 2]})

    rendered = load_workbook(output_path)
    try:
        rendered_chart = rendered.active._charts[0]
        assert isinstance(rendered_chart.anchor, OneCellAnchor)
        assert (rendered_chart.anchor._from.row, rendered_chart.anchor._from.col) == (3, 3)
        assert _chart_references(rendered_chart) == {"'Sheet'!$A$1"}
    finally:
        rendered.close()


def test_moves_chart_anchor_down_inside_cell_shift_lane(tmp_path: Path) -> None:
    template_path = tmp_path / "cell-shift-chart-template.xlsx"
    output_path = tmp_path / "cell-shift-chart-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = '{% for item in items shift="cells" %}{{ item }}'
    sheet["B1"] = "{% endfor %}"
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    sheet.add_chart(chart, "A4")
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": [1, 2, 3]})

    rendered = load_workbook(output_path)
    try:
        rendered_chart = rendered.active._charts[0]
        assert isinstance(rendered_chart.anchor, OneCellAnchor)
        assert (rendered_chart.anchor._from.row, rendered_chart.anchor._from.col) == (5, 0)
        assert _chart_references(rendered_chart) == {"'Sheet'!$A$1"}
    finally:
        rendered.close()


def test_translates_two_cell_chart_anchor_without_resizing(tmp_path: Path) -> None:
    template_path = tmp_path / "two-cell-chart-template.xlsx"
    output_path = tmp_path / "two-cell-chart-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    cast(Any, chart).anchor = TwoCellAnchor(
        _from=AnchorMarker(col=3, row=2),
        to=AnchorMarker(col=8, row=12),
    )
    sheet.add_chart(chart)
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": [1, 2]})

    rendered = load_workbook(output_path)
    try:
        rendered_chart = rendered.active._charts[0]
        assert isinstance(rendered_chart.anchor, TwoCellAnchor)
        assert (rendered_chart.anchor._from.row, rendered_chart.anchor._from.col) == (3, 3)
        assert (rendered_chart.anchor.to.row, rendered_chart.anchor.to.col) == (13, 8)
    finally:
        rendered.close()


def test_rejects_chart_anchor_copied_by_repeat(tmp_path: Path) -> None:
    template_path = tmp_path / "copied-chart-anchor-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    sheet.add_chart(chart, "A1")
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert DiagnosticCode.CHART_ANCHOR_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_rejects_two_cell_chart_anchor_resize(tmp_path: Path) -> None:
    template_path = tmp_path / "resized-chart-anchor-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A2"] = "{% for item in items %}{{ item }}"
    sheet["B2"] = "{% endfor %}"
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=2, max_row=2))
    cast(Any, chart).anchor = TwoCellAnchor(
        _from=AnchorMarker(col=3, row=0),
        to=AnchorMarker(col=8, row=3),
    )
    sheet.add_chart(chart)
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert DiagnosticCode.CHART_ANCHOR_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_preserves_distinct_styles_for_multiple_charts(tmp_path: Path) -> None:
    template_path = tmp_path / "multiple-chart-template.xlsx"
    output_path = tmp_path / "multiple-chart-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{{ value }}"
    for chart_type, style, anchor in (
        (BarChart, 10, "D2"),
        (LineChart, 11, "D20"),
    ):
        chart = chart_type()
        chart.style = style
        chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
        sheet.add_chart(chart, anchor)
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"value": 42})

    rendered = load_workbook(output_path)
    try:
        assert [type(chart) for chart in rendered.active._charts] == [BarChart, LineChart]
        assert [_chart_references(chart) for chart in rendered.active._charts] == [
            {"'Sheet'!$A$1"},
            {"'Sheet'!$A$1"},
        ]
        assert _chart_style(output_path, "chart1.xml") == "10"
        assert _chart_style(output_path, "chart2.xml") == "11"
    finally:
        rendered.close()


def test_rejects_unsupported_chart_type_and_reference(tmp_path: Path) -> None:
    template_path = tmp_path / "unsupported-chart-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Value"
    chart = DoughnutChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    chart.series[0].val.numRef.f = "DefinedChartRange"
    sheet.add_chart(chart, "C1")
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {})

    assert {
        DiagnosticCode.CHART_TYPE_UNSUPPORTED,
        DiagnosticCode.CHART_REFERENCE_UNSUPPORTED,
    } <= {diagnostic.code for diagnostic in caught.value.diagnostics}
    assert not output_path.exists()


def test_preserves_embedded_image_and_moves_it_with_row_expansion(tmp_path: Path) -> None:
    template_path = tmp_path / "image-template.xlsx"
    output_path = tmp_path / "image-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    sheet.add_image(_StaticImage(), "D3")
    _save(workbook, template_path)

    template_anchor = _drawing_anchors(template_path)[0]
    render_workbook(template_path, output_path, {"items": [1, 2]})

    output_anchor = _drawing_anchors(output_path)[0]
    assert _drawing_kind(output_anchor) == "image"
    assert _anchor_marker(template_anchor) == (2, 3)
    assert _anchor_marker(output_anchor) == (3, 3)
    assert _drawing_ext(output_anchor) == _drawing_ext(template_anchor)
    assert _picture_metadata(output_anchor) == _picture_metadata(template_anchor)
    assert _media_payloads(output_path) == _media_payloads(template_path)


def test_moves_only_images_inside_cell_shift_lane(tmp_path: Path) -> None:
    template_path = tmp_path / "lane-images-template.xlsx"
    output_path = tmp_path / "lane-images-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = '{% for item in items shift="cells" %}{{ item }}'
    sheet["B1"] = "{% endfor %}"
    sheet.add_image(_StaticImage(), "A4")
    sheet.add_image(_StaticImage(), "D4")
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": [1, 2, 3]})

    anchors = _drawing_anchors(output_path)
    assert [_drawing_kind(anchor) for anchor in anchors] == ["image", "image"]
    assert [_anchor_marker(anchor) for anchor in anchors] == [(5, 0), (3, 3)]
    assert _media_payloads(output_path) == _media_payloads(template_path)


def test_translates_two_cell_image_anchor_without_resizing(tmp_path: Path) -> None:
    template_path = tmp_path / "two-cell-image-template.xlsx"
    output_path = tmp_path / "two-cell-image-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    image = _StaticImage()
    cast(Any, image).anchor = TwoCellAnchor(
        _from=AnchorMarker(col=3, row=2),
        to=AnchorMarker(col=8, row=12),
    )
    sheet.add_image(image)
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"items": [1, 2]})

    anchor = _drawing_anchors(output_path)[0]
    assert _anchor_marker(anchor) == (3, 3)
    assert _anchor_marker(anchor, "to") == (13, 8)
    assert _media_payloads(output_path) == _media_payloads(template_path)


def test_rejects_image_anchor_copied_by_repeat(tmp_path: Path) -> None:
    template_path = tmp_path / "copied-image-anchor-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    sheet.add_image(_StaticImage(), "A1")
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert DiagnosticCode.IMAGE_ANCHOR_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_rejects_two_cell_image_anchor_resize(tmp_path: Path) -> None:
    template_path = tmp_path / "resized-image-anchor-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A2"] = "{% for item in items %}{{ item }}"
    sheet["B2"] = "{% endfor %}"
    image = _StaticImage()
    cast(Any, image).anchor = TwoCellAnchor(
        _from=AnchorMarker(col=3, row=0),
        to=AnchorMarker(col=8, row=3),
    )
    sheet.add_image(image)
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert DiagnosticCode.IMAGE_ANCHOR_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_rejects_unsupported_embedded_image_format(tmp_path: Path) -> None:
    template_path = tmp_path / "bmp-image-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    workbook.active.add_image(_StaticImage("bmp"), "C1")
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {})

    assert DiagnosticCode.IMAGE_FORMAT_UNSUPPORTED in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_preserves_embedded_jpeg_bytes(tmp_path: Path) -> None:
    template_path = tmp_path / "jpeg-image-template.xlsx"
    output_path = tmp_path / "jpeg-image-output.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{{ value }}"
    workbook.active.add_image(_StaticImage("jpeg", _ONE_PIXEL_JPEG), "C1")
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"value": 42})

    assert _media_payloads(output_path) == [_ONE_PIXEL_JPEG]
    with ZipFile(output_path) as archive:
        assert "xl/media/image1.jpeg" in archive.namelist()


def test_preserves_static_text_shape_and_moves_it_with_row_expansion(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "text-shape-template.xlsx"
    output_path = tmp_path / "text-shape-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    _save(workbook, template_path)
    workbook.close()
    _inject_text_shapes(template_path, (_shape_anchor(row=2, column=3),))
    template_anchor = _drawing_anchors(template_path)[0]

    render_workbook(template_path, output_path, {"items": ["A", "B"]})

    output_anchor = _drawing_anchors(output_path)[0]
    assert _drawing_kind(output_anchor) == "text shape"
    assert _anchor_marker(template_anchor) == (2, 3)
    assert _anchor_marker(output_anchor) == (3, 3)
    assert _drawing_ext(output_anchor) == _drawing_ext(template_anchor)
    assert _shape_text(output_anchor) == "Static {{ customer.name }}"
    assert _shape_visual_state(output_anchor) == _shape_visual_state(template_anchor)
    assert _worksheet_drawing_target(output_path) == "xl/drawings/drawing1.xml"
    rendered = load_workbook(output_path, read_only=True, data_only=False)
    rendered.close()


def test_moves_only_text_shapes_inside_cell_shift_lane(tmp_path: Path) -> None:
    template_path = tmp_path / "lane-text-shapes-template.xlsx"
    output_path = tmp_path / "lane-text-shapes-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = '{% for item in items shift="cells" %}{{ item }}'
    sheet["B1"] = "{% endfor %}"
    _save(workbook, template_path)
    workbook.close()
    _inject_text_shapes(
        template_path,
        (
            _shape_anchor(row=3, column=0, text="Inside lane"),
            _shape_anchor(row=3, column=3, text="Outside lane", preset="wedgeRectCallout"),
        ),
    )

    render_workbook(template_path, output_path, {"items": [1, 2, 3]})

    anchors = _drawing_anchors(output_path)
    assert [_drawing_kind(anchor) for anchor in anchors] == ["text shape", "text shape"]
    assert [_anchor_marker(anchor) for anchor in anchors] == [(5, 0), (3, 3)]
    assert [_shape_text(anchor) for anchor in anchors] == ["Inside lane", "Outside lane"]


def test_translates_two_cell_text_shape_without_resizing(tmp_path: Path) -> None:
    template_path = tmp_path / "two-cell-text-shape-template.xlsx"
    output_path = tmp_path / "two-cell-text-shape-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    _save(workbook, template_path)
    workbook.close()
    _inject_text_shapes(
        template_path,
        (_shape_anchor(row=2, column=3, to_row=12, to_column=8),),
    )

    render_workbook(template_path, output_path, {"items": [1, 2]})

    anchor = _drawing_anchors(output_path)[0]
    assert _anchor_marker(anchor) == (3, 3)
    assert _anchor_marker(anchor, "to") == (13, 8)


def test_rejects_text_shape_anchor_copied_by_repeat(tmp_path: Path) -> None:
    template_path = tmp_path / "copied-text-shape-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}"
    sheet["B1"] = "{% endfor %}"
    _save(workbook, template_path)
    workbook.close()
    _inject_text_shapes(template_path, (_shape_anchor(row=0, column=0),))

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert DiagnosticCode.TEXT_SHAPE_ANCHOR_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_rejects_two_cell_text_shape_resize(tmp_path: Path) -> None:
    template_path = tmp_path / "resized-text-shape-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A2"] = "{% for item in items %}{{ item }}"
    sheet["B2"] = "{% endfor %}"
    _save(workbook, template_path)
    workbook.close()
    _inject_text_shapes(
        template_path,
        (_shape_anchor(row=0, column=3, to_row=3, to_column=8),),
    )

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": [1, 2]})

    assert DiagnosticCode.TEXT_SHAPE_ANCHOR_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


@pytest.mark.parametrize("unsupported", ["field", "macro", "textlink", "wordart"])
def test_rejects_dynamic_or_linked_text_shape_content(
    tmp_path: Path,
    unsupported: str,
) -> None:
    template_path = tmp_path / f"unsupported-{unsupported}-shape.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{{ value }}"
    _save(workbook, template_path)
    workbook.close()
    _inject_text_shapes(
        template_path,
        (_shape_anchor(row=2, column=3, unsupported=unsupported),),
    )

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"value": 42})

    assert DiagnosticCode.TEXT_SHAPE_UNSUPPORTED in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_preserves_mixed_shape_image_chart_stacking_order(tmp_path: Path) -> None:
    template_path = tmp_path / "mixed-shape-drawing-template.xlsx"
    output_path = tmp_path / "mixed-shape-drawing-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{{ value }}"
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    sheet.add_chart(chart, "D2")
    sheet.add_image(_StaticImage(), "D20")
    _save(workbook, template_path)
    workbook.close()
    _inject_text_shapes(template_path, (_shape_anchor(row=9, column=3),))
    _set_drawing_order(template_path, ("image", "text shape", "chart"))

    render_workbook(template_path, output_path, {"value": 42})

    assert [_drawing_kind(anchor) for anchor in _drawing_anchors(output_path)] == [
        "image",
        "text shape",
        "chart",
    ]
    identifiers = [
        int(properties.get("id", "0"))
        for anchor in _drawing_anchors(output_path)
        for properties in anchor.iter(_qn(_DRAWING_NAMESPACE, "cNvPr"))
    ]
    assert identifiers == [1, 2, 3]
    assert _media_payloads(output_path) == _media_payloads(template_path)


def test_preserves_mixed_chart_image_stacking_order(tmp_path: Path) -> None:
    template_path = tmp_path / "mixed-drawing-template.xlsx"
    output_path = tmp_path / "mixed-drawing-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{{ value }}"
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    sheet.add_chart(chart, "D2")
    sheet.add_image(_StaticImage(), "D20")
    _save(workbook, template_path)
    _set_drawing_order(template_path, ("image", "chart"))

    render_workbook(template_path, output_path, {"value": 42})

    assert [_drawing_kind(anchor) for anchor in _drawing_anchors(template_path)] == [
        "image",
        "chart",
    ]
    assert [_drawing_kind(anchor) for anchor in _drawing_anchors(output_path)] == [
        "image",
        "chart",
    ]
    assert _media_payloads(output_path) == _media_payloads(template_path)


def test_rejects_chartsheet(tmp_path: Path) -> None:
    template_path = tmp_path / "chartsheet-template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 10
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=1))
    workbook.create_chartsheet("Chart only").add_chart(chart)
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {})

    assert [diagnostic.code for diagnostic in caught.value.diagnostics] == [
        DiagnosticCode.CHARTSHEET_UNSUPPORTED
    ]
    assert not output_path.exists()


def test_runtime_string_that_starts_with_equals_remains_text(tmp_path: Path) -> None:
    template_path = tmp_path / "formula-injection.xlsx"
    output_path = tmp_path / "formula-injection-output.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{{ supplied }}"
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {"supplied": '=HYPERLINK("bad")'})

    rendered = load_workbook(output_path, data_only=False)
    try:
        assert rendered.active["A1"].value == '=HYPERLINK("bad")'
        assert rendered.active["A1"].data_type == "s"
    finally:
        rendered.close()


def test_rejects_custom_row_height_repeated_by_cell_shift(tmp_path: Path) -> None:
    template_path = tmp_path / "cell-shift-height.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = '{% for item in items shift="cells" %}{{ item }}{% endfor %}'
    sheet.row_dimensions[1].height = 28
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": ["A", "B"]})

    assert DiagnosticCode.CELL_SHIFT_WITH_CUSTOM_ROW_HEIGHT in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    assert not output_path.exists()


def test_allows_custom_height_repeated_by_an_outer_row_shift(tmp_path: Path) -> None:
    template_path = tmp_path / "nested-height.xlsx"
    output_path = tmp_path / "nested-height-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for group in groups %}{{ group.name }}"
    sheet["A2"] = '{% for item in group.items shift="cells" %}{{ item }}{% endfor %}'
    sheet["B3"] = "{% endfor %}"
    sheet.row_dimensions[2].height = 28
    _save(workbook, template_path)

    render_workbook(
        template_path,
        output_path,
        {
            "groups": [
                {"name": "First", "items": ["A"]},
                {"name": "Second", "items": ["B"]},
            ]
        },
    )

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        assert sheet.row_dimensions[2].height == 28
        assert sheet.row_dimensions[5].height == 28
    finally:
        rendered.close()


def test_rejects_data_validation_on_a_transformed_sheet(tmp_path: Path) -> None:
    template_path = tmp_path / "validation.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "{% for item in items %}{{ item }}{% endfor %}"
    validation = DataValidation(type="list", formula1='"Open,Closed"')
    validation.add("A1:A3")
    sheet.add_data_validation(validation)
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {"items": ["A", "B"]})

    assert DiagnosticCode.DATA_VALIDATION_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }


def test_rejects_native_excel_table_even_without_layout_changes(tmp_path: Path) -> None:
    template_path = tmp_path / "table.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name"])
    sheet.append(["Alpha"])
    sheet.add_table(Table(displayName="ItemsTable", ref="A1:A2"))
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, output_path, {})

    assert DiagnosticCode.TABLE_REQUIRES_UNSUPPORTED_TRANSFORM in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }


def test_preserves_static_hyperlink_and_comment(tmp_path: Path) -> None:
    template_path = tmp_path / "links.xlsx"
    output_path = tmp_path / "links-output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "OpenAI"
    sheet["A1"].hyperlink = "https://openai.com"
    sheet["B1"] = "Reviewed"
    sheet["B1"].comment = Comment("Approved", "Reviewer")
    _save(workbook, template_path)

    render_workbook(template_path, output_path, {})

    rendered = load_workbook(output_path)
    try:
        sheet = rendered.active
        assert sheet["A1"].hyperlink.target == "https://openai.com"
        assert sheet["B1"].comment.text == "Approved"
        assert sheet["B1"].comment.author == "Reviewer"
    finally:
        rendered.close()


def test_rejects_input_path_as_output_path(tmp_path: Path) -> None:
    template_path = tmp_path / "same.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{{ value }}"
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(template_path, template_path, {"value": "Rendered"})

    assert DiagnosticCode.INPUT_OUTPUT_PATH_CONFLICT in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }
    unchanged = load_workbook(template_path)
    try:
        assert unchanged.active["A1"].value == "{{ value }}"
    finally:
        unchanged.close()


def test_rejects_non_xlsx_paths_before_loading(tmp_path: Path) -> None:
    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(tmp_path / "template.xlsm", tmp_path / "output.xlsx", {})

    assert DiagnosticCode.UNSUPPORTED_WORKBOOK_FORMAT in {
        diagnostic.code for diagnostic in caught.value.diagnostics
    }


def test_rejects_noncanonical_context_before_writing(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{% for item in items %}{{ item }}{% endfor %}"
    _save(workbook, template_path)

    with pytest.raises(TemplateRenderError) as caught:
        render_workbook(template_path, output_path, {"items": {"A", "B"}})

    assert [diagnostic.code for diagnostic in caught.value.diagnostics] == [
        DiagnosticCode.UNORDERED_CONTEXT_COLLECTION
    ]
    assert str(caught.value.diagnostics[0].location) == "context.items"
    assert not output_path.exists()


def test_rejects_input_package_over_configured_file_limit(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "Static"
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(
            template_path,
            output_path,
            {},
            limits=ResourceLimits(max_xlsx_file_bytes=1),
        )

    assert [diagnostic.code for diagnostic in caught.value.diagnostics] == [
        DiagnosticCode.XLSX_PACKAGE_LIMIT_EXCEEDED
    ]
    assert not output_path.exists()


def test_rejects_input_package_over_configured_member_limit(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(
            template_path,
            output_path,
            {},
            limits=ResourceLimits(max_xlsx_archive_members=1),
        )

    assert [diagnostic.code for diagnostic in caught.value.diagnostics] == [
        DiagnosticCode.XLSX_PACKAGE_LIMIT_EXCEEDED
    ]
    assert not output_path.exists()


def test_rejects_workbook_over_configured_sheet_limit(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    workbook.create_sheet("Second")
    _save(workbook, template_path)

    with pytest.raises(TemplateCompilationError) as caught:
        render_workbook(
            template_path,
            output_path,
            {},
            limits=ResourceLimits(max_worksheets=1),
        )

    assert [diagnostic.code for diagnostic in caught.value.diagnostics] == [
        DiagnosticCode.WORKSHEET_COUNT_LIMIT_EXCEEDED
    ]
    assert not output_path.exists()


def test_rejects_workbook_wide_planned_cell_limit(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "First"
    workbook.create_sheet("Second")["A1"] = "Second"
    _save(workbook, template_path)

    with pytest.raises(TemplateRenderError) as caught:
        render_workbook(
            template_path,
            output_path,
            {},
            limits=ResourceLimits(max_planned_cells_per_workbook=1),
        )

    assert [diagnostic.code for diagnostic in caught.value.diagnostics] == [
        DiagnosticCode.RENDER_RESOURCE_LIMIT_EXCEEDED
    ]
    assert not output_path.exists()


def test_rejects_oversized_temporary_output_before_publication(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "should-not-exist.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "{% for item in items %}{{ item }}{% endfor %}"
    _save(workbook, template_path)
    limits = replace(
        ResourceLimits(),
        max_xlsx_uncompressed_bytes=_uncompressed_size(template_path) + 5_000,
    )
    values = [f"row-{index:03d}-" + ("x" * 500) for index in range(100)]

    with pytest.raises(TemplateRenderError) as caught:
        render_workbook(template_path, output_path, {"items": values}, limits=limits)

    assert [diagnostic.code for diagnostic in caught.value.diagnostics] == [
        DiagnosticCode.XLSX_PACKAGE_LIMIT_EXCEEDED
    ]
    assert not output_path.exists()
